"""Generate demo XLSX files for testing ream-xlsx.

Run once to create the demo workbooks:
    python demo/generate_demos.py
"""

from datetime import date
from pathlib import Path

import openpyxl

DEMO_DIR = Path(__file__).parent


def create_simple_table():
    """A basic employee table — the simplest possible input."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Employees"

    data = [
        ["Name", "Department", "Salary", "Start Date"],
        ["Alice", "Engineering", 120000, date(2021, 3, 15)],
        ["Bob", "Marketing", 95000, date(2022, 7, 1)],
        ["Carol", "Engineering", 135000, date(2020, 1, 10)],
        ["Dave", "Sales", 88000, date(2023, 5, 20)],
        ["Eve", "Marketing", 102000, date(2021, 11, 8)],
    ]
    for row in data:
        ws.append(row)

    wb.save(DEMO_DIR / "simple_table.xlsx")
    print("Created simple_table.xlsx")


def create_multi_sheet():
    """A workbook with multiple sheets — tests #!SHEET directives."""
    wb = openpyxl.Workbook()

    # Sheet 1: Revenue
    ws1 = wb.active
    ws1.title = "Revenue"
    months = ["", "Jan", "Feb", "Mar", "Q1 Total"]
    ws1.append(months)
    ws1.append(["Product A", 10000, 12000, 11000, 33000])
    ws1.append(["Product B", 8000, 8500, 9200, 25700])
    ws1.append(["Product C", 15000, 14000, 16000, 45000])
    ws1.append(["Total", 33000, 34500, 36200, 103700])

    # Sheet 2: Expenses
    ws2 = wb.create_sheet("Expenses")
    ws2.append(["Category", "Jan", "Feb", "Mar", "Q1 Total"])
    ws2.append(["Payroll", 25000, 25000, 25000, 75000])
    ws2.append(["Rent", 5000, 5000, 5000, 15000])
    ws2.append(["Software", 2000, 2000, 2000, 6000])
    ws2.append(["Travel", 1500, 3000, 800, 5300])
    ws2.append(["Total", 33500, 35000, 32800, 101300])

    # Sheet 3: Summary
    ws3 = wb.create_sheet("Summary")
    ws3.append(["Metric", "Q1"])
    ws3.append(["Revenue", 103700])
    ws3.append(["Expenses", 101300])
    ws3.append(["Net Income", 2400])

    wb.save(DEMO_DIR / "multi_sheet.xlsx")
    print("Created multi_sheet.xlsx")


def create_sparse_data():
    """A sheet with gaps in rows and columns — tests sparse encoding."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sensor Readings"

    ws.append(["Sensor ID", "Location", None, "Status", None, "Last Reading"])

    # Row 2
    ws.cell(row=2, column=1, value="S-001")
    ws.cell(row=2, column=2, value="Building A")
    ws.cell(row=2, column=4, value="Active")
    ws.cell(row=2, column=6, value=72.4)

    # Row 3 — skip row intentionally
    # Row 4
    ws.cell(row=4, column=1, value="S-002")
    ws.cell(row=4, column=2, value="Building B")
    ws.cell(row=4, column=4, value="Inactive")
    ws.cell(row=4, column=6, value=None)

    # Rows 5-6 — skip
    # Row 7
    ws.cell(row=7, column=1, value="S-003")
    ws.cell(row=7, column=2, value="Building A")
    ws.cell(row=7, column=4, value="Active")
    ws.cell(row=7, column=6, value=68.1)

    # Row 10 — big gap
    ws.cell(row=10, column=1, value="S-004")
    ws.cell(row=10, column=2, value="Building C")
    ws.cell(row=10, column=4, value="Active")
    ws.cell(row=10, column=6, value=71.9)

    wb.save(DEMO_DIR / "sparse_data.xlsx")
    print("Created sparse_data.xlsx")


def create_repeated_rows():
    """A sheet with repeated identical rows — tests collapse_rows option."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Pricing"

    ws.append(["Region", "Plan", "Monthly Price", "Annual Price"])

    # 5 identical rows for US Standard
    for _ in range(5):
        ws.append(["US", "Standard", 29, 290])

    # 3 identical rows for US Premium
    for _ in range(3):
        ws.append(["US", "Premium", 59, 590])

    # 5 identical rows for EU Standard
    for _ in range(5):
        ws.append(["EU", "Standard", 29, 290])

    # 3 identical rows for EU Premium
    for _ in range(3):
        ws.append(["EU", "Premium", 59, 590])

    wb.save(DEMO_DIR / "repeated_rows.xlsx")
    print("Created repeated_rows.xlsx")


def create_financial_model():
    """A mini P&L with formulas — tests real-world financial spreadsheets."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "P&L"

    # Headers
    ws.append(["", "2023", "2024", "2025 (proj)"])

    # Revenue section
    ws.append(["Revenue", 500000, 620000, 780000])
    ws.append(["COGS", 200000, 240000, 295000])
    ws.append(["Gross Profit", 300000, 380000, 485000])
    ws.append([])  # blank row

    # Operating expenses
    ws.append(["Operating Expenses"])
    ws.append(["  Salaries", 150000, 180000, 220000])
    ws.append(["  Rent", 36000, 36000, 42000])
    ws.append(["  Marketing", 25000, 40000, 55000])
    ws.append(["  Other", 12000, 15000, 18000])
    ws.append(["Total OpEx", 223000, 271000, 335000])
    ws.append([])  # blank row

    # Bottom line
    ws.append(["EBITDA", 77000, 109000, 150000])
    ws.append(["Tax (25%)", 19250, 27250, 37500])
    ws.append(["Net Income", 57750, 81750, 112500])

    # Assumptions sheet
    ws2 = wb.create_sheet("Assumptions")
    ws2.append(["Parameter", "Value"])
    ws2.append(["Tax Rate", 0.25])
    ws2.append(["Revenue Growth", 0.26])
    ws2.append(["COGS %", 0.38])
    ws2.append(["Headcount", 12])

    wb.save(DEMO_DIR / "financial_model.xlsx")
    print("Created financial_model.xlsx")


def create_mixed_types():
    """A sheet with various data types — booleans, dates, numbers, strings."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Mixed Data"

    ws.append(["Field", "Value", "Type"])
    ws.append(["Name", "Acme Corp", "string"])
    ws.append(["Founded", date(1998, 6, 15), "date"])
    ws.append(["Revenue", 1250000.50, "float"])
    ws.append(["Employees", 47, "integer"])
    ws.append(["Public", True, "boolean"])
    ws.append(["Profitable", False, "boolean"])
    ws.append(["Notes", "Has a pipe | in name", "string with special char"])
    ws.append(["Ticker", "", "empty string"])
    ws.append(["Growth", 0.156, "percentage"])

    wb.save(DEMO_DIR / "mixed_types.xlsx")
    print("Created mixed_types.xlsx")


if __name__ == "__main__":
    create_simple_table()
    create_multi_sheet()
    create_sparse_data()
    create_repeated_rows()
    create_financial_model()
    create_mixed_types()
    print("\nAll demo files created in demo/")
