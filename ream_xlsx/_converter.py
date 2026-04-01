"""Core REAM conversion logic -- ported from src/converters.py."""

from __future__ import annotations

import re
from datetime import date, datetime
from typing import Any

from openpyxl.utils import get_column_letter
from openpyxl.workbook.workbook import Workbook

from ream_xlsx._formula import _a1_formula_to_r1c1
from ream_xlsx._options import ReamOptions


def _cell_value_str(val: Any) -> str:
    """Convert a cell value to a string representation."""
    if val is None:
        return ""
    if isinstance(val, bool):
        return "TRUE" if val else "FALSE"
    if isinstance(val, datetime):
        return val.strftime("%Y-%m-%d")
    if isinstance(val, date):
        return val.strftime("%Y-%m-%d")
    if isinstance(val, float):
        # Remove trailing zeros
        if val == int(val) and abs(val) < 1e15:
            return str(int(val))
        return f"{val:.10g}"
    return str(val)


def _needs_ream_quoting(s: str) -> bool:
    """Check if a string needs quoting in Ream format."""
    if not s:
        return True  # empty string
    if s.startswith("=") or s.startswith("[") or s.startswith('"') or s.startswith("@"):
        return True
    if "|" in s or '"' in s or "\n" in s or "\r" in s:
        return True
    if s != s.strip():
        return True  # leading/trailing whitespace
    if s in ("TRUE", "FALSE", "true", "false"):
        return True
    # Check if it looks like a number
    try:
        float(s)
        return True
    except ValueError:
        pass
    # Check if it looks like an error
    if s in ("#N/A", "#REF!", "#VALUE!", "#DIV/0!", "#NULL!", "#NAME?", "#NUM!"):
        return True
    # Check if it matches A1-style addressed-entry pattern (e.g. B=, AB=, B:M=)
    if re.match(r"^[A-Z]{1,3}(?::[A-Z]{1,3})?=", s):
        return True
    return False


def _ream_quote(s: str) -> str:
    """Quote a string for Ream format."""
    escaped = s.replace("\\", "\\\\").replace('"', '""').replace("\n", "\\n").replace("\r", "\\r").replace("\t", "\\t")
    return f'"{escaped}"'


def _ream_scalar(val: Any) -> str:
    """Convert a cell value to a Ream scalar."""
    if val is None:
        return ""
    if isinstance(val, bool):
        return "TRUE" if val else "FALSE"
    if isinstance(val, (int, float)):
        if isinstance(val, float):
            if val == int(val) and abs(val) < 1e15:
                return str(int(val))
            return f"{val:.10g}"
        return str(val)
    if isinstance(val, datetime):
        s = val.strftime("%Y-%m-%d")
        return s
    if isinstance(val, date):
        s = val.strftime("%Y-%m-%d")
        return s
    s = str(val)
    if _needs_ream_quoting(s):
        return _ream_quote(s)
    return s


def _cell_to_ream(val: Any, row: int, col: int, emit_formulas: bool) -> tuple[str, bool]:
    """Convert a cell value to its REAM representation.

    Returns:
        (ream_text, is_formula) — *ream_text* is the R1C1 formula body
        (without leading ``=``) when *is_formula* is True, or the normal
        scalar text otherwise.
    """
    if emit_formulas and isinstance(val, str) and val.startswith("="):
        formula_body = val[1:]  # strip the leading '='
        r1c1_body = _a1_formula_to_r1c1(formula_body, row, col)
        return r1c1_body, True
    return _ream_scalar(val), False


def _xlsx_to_ream_impl(wb: Workbook, options: ReamOptions) -> str:
    """Convert an already-loaded Workbook to REAM text.

    Args:
        wb: An openpyxl Workbook.
        options: Conversion options.

    Returns:
        REAM-formatted string.
    """
    version = "11" if options.collapse_rows else "9"
    lines = [f"#!REAM {version}"]

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        safe_name = sheet_name
        if " " in sheet_name or "|" in sheet_name or '"' in sheet_name or "&" in sheet_name:
            safe_name = _ream_quote(sheet_name)
        lines.append(f"#!SHEET {safe_name}")

        header_row = None
        for row_idx in range(1, min(ws.max_row or 1, options.max_rows_per_sheet) + 1):
            for col_idx in range(1, (ws.max_column or 1) + 1):
                if ws.cell(row=row_idx, column=col_idx).value is not None:
                    header_row = row_idx
                    break
            if header_row:
                break

        if header_row:
            lines.append(f"#!HEADERS {header_row}:{header_row}")

        # Collect all row data
        row_segments: dict[int, Any] = {}
        row_count = 0
        for row_idx in range(1, (ws.max_row or 0) + 1):
            if row_count >= options.max_rows_per_sheet:
                lines.append(f"# ... showing first {options.max_rows_per_sheet} data rows")
                break

            cells: dict[int, str] = {}
            cell_is_formula: dict[int, bool] = {}
            for col_idx in range(1, (ws.max_column or 0) + 1):
                val = ws.cell(row=row_idx, column=col_idx).value
                if val is not None:
                    text, is_formula = _cell_to_ream(val, row_idx, col_idx, options.emit_formulas)
                    cells[col_idx] = text
                    cell_is_formula[col_idx] = is_formula

            if not cells:
                continue

            row_count += 1

            if options.collapse_rows:
                # Build maximal horizontal segments of identical values
                segments: list[tuple[int, int, str, bool]] = []
                sorted_cols = sorted(cells.keys())
                i = 0
                while i < len(sorted_cols):
                    start_col = sorted_cols[i]
                    cell_val = cells[start_col]
                    is_f = cell_is_formula.get(start_col, False)
                    end_col = start_col
                    while (
                        i + 1 < len(sorted_cols)
                        and sorted_cols[i + 1] == end_col + 1
                        and cells[sorted_cols[i + 1]] == cell_val
                        and cell_is_formula.get(sorted_cols[i + 1], False) == is_f
                    ):
                        i += 1
                        end_col = sorted_cols[i]
                    segments.append((start_col, end_col, cell_val, is_f))
                    i += 1
                row_segments[row_idx] = segments
            else:
                # Store raw cells for simple per-row emission
                row_segments[row_idx] = (cells, cell_is_formula)

        if options.collapse_rows:
            # Vertical merge: group consecutive rows with identical segment lists
            sorted_rows = sorted(row_segments.keys())
            merged_records: list[tuple[int, int, list[tuple[int, int, str, bool]]]] = []
            i = 0
            while i < len(sorted_rows):
                start_row = sorted_rows[i]
                segs = row_segments[start_row]
                end_row = start_row
                while i + 1 < len(sorted_rows) and sorted_rows[i + 1] == end_row + 1:
                    if row_segments[sorted_rows[i + 1]] == segs:
                        end_row = sorted_rows[i + 1]
                        i += 1
                    else:
                        break
                merged_records.append((start_row, end_row, segs))
                i += 1

            for start_row, end_row, segs in merged_records:
                entries: list[str] = []
                cursor = 1
                for start_col, end_col, val_str, is_f in segs:
                    col_start = get_column_letter(start_col)
                    col_end = get_column_letter(end_col)
                    if options.force_col_selectors:
                        if start_col == end_col:
                            # Addressed entry: COL=body (formula body has no leading =)
                            entries.append(f"{col_start}={val_str}")
                        else:
                            entries.append(f"{col_start}:{col_end}={val_str}")
                    else:
                        if start_col == cursor and start_col == end_col:
                            # Bare entry: formulas need '=' prefix
                            if is_f:
                                entries.append(f"={val_str}")
                            else:
                                entries.append(val_str)
                        elif start_col == end_col:
                            entries.append(f"{col_start}={val_str}")
                        else:
                            entries.append(f"{col_start}:{col_end}={val_str}")
                    cursor = end_col + 1

                prefix = str(start_row) if start_row == end_row else f"{start_row}:{end_row}"
                lines.append(f"{prefix} | " + " | ".join(entries) + " |")
        else:
            # Simple per-row emission
            for row_idx in sorted(row_segments.keys()):
                row_cells, row_formulas = row_segments[row_idx]
                entries = []
                cursor = 1
                for col_idx in sorted(row_cells.keys()):
                    val_str = row_cells[col_idx]
                    is_f = row_formulas.get(col_idx, False)
                    col_letter = get_column_letter(col_idx)
                    if options.force_col_selectors:
                        # Addressed entry: formula body goes after COL=
                        entries.append(f"{col_letter}={val_str}")
                    else:
                        if col_idx == cursor:
                            # Bare entry: formulas need '=' prefix
                            if is_f:
                                entries.append(f"={val_str}")
                            else:
                                entries.append(val_str)
                        else:
                            # Addressed entry
                            entries.append(f"{col_letter}={val_str}")
                    cursor = col_idx + 1
                lines.append(f"{row_idx} | " + " | ".join(entries) + " |")

    return "\n".join(lines)
