"""TDD tests for the public REAM API: xlsx_to_ream, bytes_to_ream, file_to_ream."""

from __future__ import annotations

from io import BytesIO
from pathlib import Path

import openpyxl
import pytest
from openpyxl import Workbook

from ream_xlsx import (
    ConversionError,
    ReamOptions,
    bytes_to_ream,
    file_to_ream,
    xlsx_to_ream,
)


# ---------------------------------------------------------------------------
# Fixture helpers
# ---------------------------------------------------------------------------


def _make_simple_workbook() -> Workbook:
    """1 sheet, header row + 2 data rows with mixed types."""
    wb = Workbook()
    ws = wb.active
    assert ws is not None
    ws.title = "Sheet1"
    ws.append(["Name", "Revenue", "Active"])
    ws.append(["Alpha", 1000, True])
    ws.append(["Beta", 2500.5, False])
    return wb


def _make_multi_sheet_workbook() -> Workbook:
    """2 sheets with different data."""
    wb = Workbook()
    ws1 = wb.active
    assert ws1 is not None
    ws1.title = "Sales"
    ws1.append(["Region", "Amount"])
    ws1.append(["East", 100])

    ws2 = wb.create_sheet("Inventory")
    ws2.append(["Item", "Count"])
    ws2.append(["Widgets", 50])
    return wb


def _make_large_workbook() -> Workbook:
    """1 sheet with 10+ data rows for max_rows testing."""
    wb = Workbook()
    ws = wb.active
    assert ws is not None
    ws.title = "Data"
    ws.append(["ID", "Value"])
    for i in range(1, 15):
        ws.append([i, i * 10])
    return wb


def _workbook_to_bytes(wb: Workbook) -> bytes:
    """Serialize workbook to bytes via BytesIO."""
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


def _workbook_to_path(wb: Workbook, tmp_path: Path) -> str:
    """Save workbook to tmp_path and return string path."""
    p = tmp_path / "workbook.xlsx"
    wb.save(p)
    return str(p)


# ---------------------------------------------------------------------------
# Tests
# ---------------------------------------------------------------------------


def test_xlsx_to_ream_returns_ream_string(tmp_path: Path) -> None:
    """xlsx_to_ream returns a string starting with '#!REAM 9'."""
    wb = _make_simple_workbook()
    path = _workbook_to_path(wb, tmp_path)
    result = xlsx_to_ream(path)
    assert isinstance(result, str)
    assert result.startswith("#!REAM 9")


def test_xlsx_to_ream_contains_sheet_header(tmp_path: Path) -> None:
    """Output contains '#!SHEET Sheet1' for a sheet named Sheet1."""
    wb = _make_simple_workbook()
    path = _workbook_to_path(wb, tmp_path)
    result = xlsx_to_ream(path)
    assert "#!SHEET Sheet1" in result


def test_xlsx_to_ream_contains_headers_directive(tmp_path: Path) -> None:
    """Output contains '#!HEADERS 1:1' for a workbook with a header row."""
    wb = _make_simple_workbook()
    path = _workbook_to_path(wb, tmp_path)
    result = xlsx_to_ream(path)
    assert "#!HEADERS 1:1" in result


def test_xlsx_to_ream_contains_data_rows(tmp_path: Path) -> None:
    """Output contains row data with cell values from the workbook."""
    wb = _make_simple_workbook()
    path = _workbook_to_path(wb, tmp_path)
    result = xlsx_to_ream(path)
    assert "Alpha" in result
    assert "1000" in result


def test_bytes_to_ream_returns_ream_string() -> None:
    """bytes_to_ream returns a string starting with '#!REAM'."""
    wb = _make_simple_workbook()
    data = _workbook_to_bytes(wb)
    result = bytes_to_ream(data)
    assert isinstance(result, str)
    assert result.startswith("#!REAM")


def test_file_to_ream_returns_ream_string() -> None:
    """file_to_ream returns a string starting with '#!REAM'."""
    wb = _make_simple_workbook()
    stream = BytesIO(_workbook_to_bytes(wb))
    result = file_to_ream(stream)
    assert isinstance(result, str)
    assert result.startswith("#!REAM")


def test_bytes_to_ream_matches_path(tmp_path: Path) -> None:
    """bytes_to_ream output matches xlsx_to_ream for the same workbook."""
    wb = _make_simple_workbook()
    path = _workbook_to_path(wb, tmp_path)
    # Reload from disk to get identical bytes
    wb2 = openpyxl.load_workbook(path)
    data = _workbook_to_bytes(wb2)
    result_path = xlsx_to_ream(path)
    result_bytes = bytes_to_ream(data)
    assert result_path == result_bytes


def test_file_to_ream_matches_path(tmp_path: Path) -> None:
    """file_to_ream output matches xlsx_to_ream for the same workbook."""
    wb = _make_simple_workbook()
    path = _workbook_to_path(wb, tmp_path)
    wb2 = openpyxl.load_workbook(path)
    stream = BytesIO(_workbook_to_bytes(wb2))
    result_path = xlsx_to_ream(path)
    result_stream = file_to_ream(stream)
    assert result_path == result_stream


def test_xlsx_to_ream_accepts_path_object(tmp_path: Path) -> None:
    """xlsx_to_ream accepts a pathlib.Path argument (API-04)."""
    wb = _make_simple_workbook()
    p = tmp_path / "workbook.xlsx"
    wb.save(p)
    result = xlsx_to_ream(p)
    assert result.startswith("#!REAM")


def test_output_is_deterministic(tmp_path: Path) -> None:
    """xlsx_to_ream called twice returns byte-for-byte identical output (API-06)."""
    wb = _make_simple_workbook()
    path = _workbook_to_path(wb, tmp_path)
    result1 = xlsx_to_ream(path)
    result2 = xlsx_to_ream(path)
    assert result1 == result2


def test_max_rows_per_sheet(tmp_path: Path) -> None:
    """ReamOptions(max_rows_per_sheet=1) limits data rows to 1 per sheet."""
    wb = _make_large_workbook()
    path = _workbook_to_path(wb, tmp_path)
    opts = ReamOptions(max_rows_per_sheet=1)
    result = xlsx_to_ream(path, opts)
    lines = result.splitlines()
    # Count data row lines (non-directive lines)
    data_lines = [l for l in lines if l and not l.startswith("#")]
    assert len(data_lines) <= 1


def test_collapse_rows_version(tmp_path: Path) -> None:
    """ReamOptions(collapse_rows=True) produces '#!REAM 11' header."""
    wb = _make_simple_workbook()
    path = _workbook_to_path(wb, tmp_path)
    opts = ReamOptions(collapse_rows=True)
    result = xlsx_to_ream(path, opts)
    assert result.startswith("#!REAM 11")


def test_force_col_selectors(tmp_path: Path) -> None:
    """ReamOptions(force_col_selectors=True) prefixes every cell with column letter."""
    wb = _make_simple_workbook()
    path = _workbook_to_path(wb, tmp_path)
    opts = ReamOptions(force_col_selectors=True)
    result = xlsx_to_ream(path, opts)
    # Every data row should have at least one 'X=value' entry with column prefix
    lines = result.splitlines()
    data_lines = [l for l in lines if l and not l.startswith("#")]
    assert len(data_lines) > 0
    for line in data_lines:
        # At least the first column should have a prefix like 'A=...'
        assert "A=" in line


def test_multi_sheet_workbook(tmp_path: Path) -> None:
    """Workbook with 2+ sheets produces 2+ #!SHEET directives."""
    wb = _make_multi_sheet_workbook()
    path = _workbook_to_path(wb, tmp_path)
    result = xlsx_to_ream(path)
    sheet_count = result.count("#!SHEET")
    assert sheet_count >= 2


def test_empty_sheet_handling(tmp_path: Path) -> None:
    """Sheet with no data does not crash."""
    wb = Workbook()
    ws = wb.active
    assert ws is not None
    ws.title = "Empty"
    path = _workbook_to_path(wb, tmp_path)
    result = xlsx_to_ream(path)
    assert "#!SHEET Empty" in result


def test_conversion_error_on_failure() -> None:
    """ConversionError is raised if conversion logic fails internally."""
    # We'll trigger this by patching _xlsx_to_ream_impl to raise an unexpected exception
    import unittest.mock as mock

    from ream_xlsx import _io  # noqa: F401

    with mock.patch(
        "ream_xlsx._converter._xlsx_to_ream_impl",
        side_effect=RuntimeError("internal boom"),
    ):
        with pytest.raises(ConversionError):
            wb = _make_simple_workbook()
            buf = BytesIO()
            wb.save(buf)
            buf.seek(0)
            bytes_to_ream(buf.read())
