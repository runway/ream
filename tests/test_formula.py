"""Tests for formula preservation and A1-to-R1C1 conversion."""

from __future__ import annotations

from io import BytesIO
from pathlib import Path

import pytest
from openpyxl import Workbook

from ream_xlsx import ReamOptions, bytes_to_ream, file_to_ream, xlsx_to_ream
from ream_xlsx._formula import _a1_formula_to_r1c1, _col_to_num


# ---------------------------------------------------------------------------
# _col_to_num
# ---------------------------------------------------------------------------


@pytest.mark.parametrize(
    ("col", "expected"),
    [("A", 1), ("B", 2), ("Z", 26), ("AA", 27), ("AZ", 52), ("XFD", 16384)],
)
def test_col_to_num(col: str, expected: int) -> None:
    assert _col_to_num(col) == expected


# ---------------------------------------------------------------------------
# _a1_formula_to_r1c1  — unit tests
# ---------------------------------------------------------------------------


def test_simple_relative_ref() -> None:
    """A1 from cell B2 → R[-1]C[-1]."""
    assert _a1_formula_to_r1c1("A1+1", 2, 2) == "R[-1]C[-1]+1"


def test_same_cell_ref() -> None:
    """A1 from cell A1 → RC (zero offsets omitted)."""
    assert _a1_formula_to_r1c1("A1", 1, 1) == "RC"


def test_absolute_ref() -> None:
    """$A$1 → R1C1 regardless of cell position."""
    assert _a1_formula_to_r1c1("$A$1", 5, 5) == "R1C1"


def test_mixed_ref_abs_col() -> None:
    """$A1 from row 3 → R[-2]C1."""
    assert _a1_formula_to_r1c1("$A1", 3, 2) == "R[-2]C1"


def test_mixed_ref_abs_row() -> None:
    """A$1 from col 3 → R1C[-2]."""
    assert _a1_formula_to_r1c1("A$1", 5, 3) == "R1C[-2]"


def test_range_ref() -> None:
    """A1:B2 from C3 → R[-2]C[-2]:R[-1]C[-1]."""
    result = _a1_formula_to_r1c1("A1:B2", 3, 3)
    assert result == "R[-2]C[-2]:R[-1]C[-1]"


def test_function_not_matched() -> None:
    """SUM( should not be treated as a cell reference."""
    result = _a1_formula_to_r1c1("SUM(A1:A10)", 1, 2)
    assert result.startswith("SUM(")
    assert "R" in result  # A1 and A10 should be converted


def test_sum_formula() -> None:
    """Full SUM formula conversion."""
    result = _a1_formula_to_r1c1("SUM(A1:A10)", 11, 1)
    assert result == "SUM(R[-10]C:R[-1]C)"


def test_string_literal_preserved() -> None:
    """References inside string literals are not converted."""
    result = _a1_formula_to_r1c1('"Cell A1 ref"', 1, 1)
    assert result == '"Cell A1 ref"'


def test_sheet_qualified_ref() -> None:
    """Sheet1!A1 from B2 → Sheet1!R[-1]C[-1]."""
    result = _a1_formula_to_r1c1("Sheet1!A1", 2, 2)
    assert result == "Sheet1!R[-1]C[-1]"


def test_quoted_sheet_ref() -> None:
    """'My Sheet'!A1 from B2 → 'My Sheet'!R[-1]C[-1]."""
    result = _a1_formula_to_r1c1("'My Sheet'!A1", 2, 2)
    assert result == "'My Sheet'!R[-1]C[-1]"


def test_complex_formula() -> None:
    """IF with mixed refs."""
    result = _a1_formula_to_r1c1("IF(A1>0,B1/A1,0)", 1, 3)
    assert result == "IF(R[-0]C[-2]>0,RC[-1]/R[-0]C[-2],0)" or result == "IF(RC[-2]>0,RC[-1]/RC[-2],0)"


def test_named_range_preserved() -> None:
    """Named ranges (lowercase, underscores) are not converted."""
    result = _a1_formula_to_r1c1("tax_rate*B2", 2, 3)
    assert "tax_rate" in result
    assert "RC[-1]" in result


def test_positive_offset() -> None:
    """A5 from A1 → R[4]C."""
    result = _a1_formula_to_r1c1("A5", 1, 1)
    assert result == "R[4]C"


# ---------------------------------------------------------------------------
# Integration: formula cells in workbook output
# ---------------------------------------------------------------------------


def _make_formula_workbook() -> Workbook:
    """Create a workbook with formula cells."""
    wb = Workbook()
    ws = wb.active
    assert ws is not None
    ws.title = "Sheet1"
    ws["A1"] = 10
    ws["A2"] = 20
    ws["A3"] = "=A1+A2"  # formula
    ws["B1"] = 100
    ws["B2"] = "=B1*2"  # formula
    return wb


def _wb_to_bytes(wb: Workbook) -> bytes:
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _wb_to_path(wb: Workbook, tmp_path: Path) -> str:
    p = tmp_path / "formulas.xlsx"
    wb.save(p)
    return str(p)


def test_formulas_preserved_in_output(tmp_path: Path) -> None:
    """Formulas appear as R1C1 in REAM output by default."""
    wb = _make_formula_workbook()
    path = _wb_to_path(wb, tmp_path)
    result = xlsx_to_ream(path)
    # A3 has =A1+A2, from row 3 col 1: A1=R[-2]C, A2=R[-1]C
    assert "R[-2]C+R[-1]C" in result


def test_formulas_preserved_bytes() -> None:
    """bytes_to_ream also preserves formulas."""
    wb = _make_formula_workbook()
    result = bytes_to_ream(_wb_to_bytes(wb))
    assert "R[-2]C+R[-1]C" in result


def test_formulas_preserved_stream() -> None:
    """file_to_ream also preserves formulas."""
    wb = _make_formula_workbook()
    result = file_to_ream(BytesIO(_wb_to_bytes(wb)))
    assert "R[-2]C+R[-1]C" in result


def test_formula_bare_entry_has_equals_prefix(tmp_path: Path) -> None:
    """Bare formula entries start with '=' per REAM spec."""
    wb = _make_formula_workbook()
    path = _wb_to_path(wb, tmp_path)
    result = xlsx_to_ream(path)
    lines = result.splitlines()
    # Row 3 has formula in col A (bare position)
    row3 = [ln for ln in lines if ln.startswith("3 |")]
    assert len(row3) == 1
    # The bare formula should start with = (e.g. "=R[-2]C+R[-1]C")
    assert "| =R[-2]C+R[-1]C |" in row3[0]


def test_formula_addressed_entry(tmp_path: Path) -> None:
    """Addressed formula entries use COL=body (no extra =)."""
    wb = _make_formula_workbook()
    path = _wb_to_path(wb, tmp_path)
    opts = ReamOptions(force_col_selectors=True)
    result = xlsx_to_ream(path, opts)
    # B2 has =B1*2, from row 2 col 2: B1=R[-1]C → formula body is R[-1]C*2
    # Addressed form: B=R[-1]C*2
    assert "B=R[-1]C*2" in result


def test_values_only_mode(tmp_path: Path) -> None:
    """emit_formulas=False falls back to cached values (no formulas)."""
    wb = _make_formula_workbook()
    path = _wb_to_path(wb, tmp_path)
    opts = ReamOptions(emit_formulas=False)
    result = xlsx_to_ream(path, opts)
    # Should NOT contain R1C1 formula references
    assert "R[-2]C" not in result
    assert "R[-1]C" not in result


def test_formula_with_collapse_rows(tmp_path: Path) -> None:
    """Formulas work correctly with collapse_rows=True."""
    wb = Workbook()
    ws = wb.active
    assert ws is not None
    ws["A1"] = "Header"
    ws["A2"] = "=A1"
    ws["A3"] = "=A2"
    path = _wb_to_path(wb, tmp_path)
    opts = ReamOptions(collapse_rows=True)
    result = xlsx_to_ream(path, opts)
    assert "#!REAM 11" in result
    # Formulas should be present
    assert "=R[-1]C" in result
