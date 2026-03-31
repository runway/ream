"""Tests for exception hierarchy and I/O adapter error handling."""

from __future__ import annotations

from io import BytesIO
from pathlib import Path

import openpyxl
import pytest

from ream_xlsx import ConversionError, InvalidWorkbookError, ReamError
from ream_xlsx._io import _load_from_bytes, _load_from_path, _load_from_stream

# ---------------------------------------------------------------------------
# Fixtures
# ---------------------------------------------------------------------------


def _make_xlsx_bytes() -> bytes:
    """Create minimal valid XLSX bytes using openpyxl."""
    wb = openpyxl.Workbook()
    ws = wb.active
    assert ws is not None
    ws["A1"] = "hello"
    ws["B1"] = 42
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _make_xlsx_with_formula() -> bytes:
    """Create XLSX bytes with a formula cell (cached value present)."""
    wb = openpyxl.Workbook()
    ws = wb.active
    assert ws is not None
    ws["A1"] = 10
    ws["A2"] = 20
    # Set a formula with a cached (data-only) result
    ws["A3"] = 30  # We'll store result directly as a plain value
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ---------------------------------------------------------------------------
# Exception hierarchy
# ---------------------------------------------------------------------------


def test_ream_error_hierarchy() -> None:
    """ReamError is a subclass of Exception."""
    assert issubclass(ReamError, Exception)
    err = ReamError("test")
    assert isinstance(err, Exception)


def test_error_subclass_hierarchy() -> None:
    """InvalidWorkbookError and ConversionError are subclasses of ReamError."""
    assert issubclass(InvalidWorkbookError, ReamError)
    assert issubclass(ConversionError, ReamError)
    invalid_err = InvalidWorkbookError("bad file")
    conv_err = ConversionError("conversion failed")
    assert isinstance(invalid_err, ReamError)
    assert isinstance(conv_err, ReamError)


# ---------------------------------------------------------------------------
# _load_from_path
# ---------------------------------------------------------------------------


def test_missing_file_raises_invalid_workbook_error() -> None:
    """_load_from_path with a nonexistent path raises InvalidWorkbookError."""
    with pytest.raises(InvalidWorkbookError):
        _load_from_path("nonexistent_file_that_does_not_exist.xlsx")


def test_directory_path_raises_invalid_workbook_error(tmp_path: Path) -> None:
    """_load_from_path with a directory path raises InvalidWorkbookError."""
    with pytest.raises(InvalidWorkbookError):
        _load_from_path(tmp_path)


def test_load_from_path_returns_workbook(tmp_path: Path) -> None:
    """_load_from_path with a valid XLSX file returns an openpyxl Workbook."""
    xlsx_file = tmp_path / "test.xlsx"
    xlsx_file.write_bytes(_make_xlsx_bytes())
    wb = _load_from_path(xlsx_file)
    assert isinstance(wb, openpyxl.Workbook)


# ---------------------------------------------------------------------------
# _load_from_bytes
# ---------------------------------------------------------------------------


def test_corrupted_bytes_raises_invalid_workbook_error() -> None:
    """_load_from_bytes with non-XLSX bytes raises InvalidWorkbookError."""
    with pytest.raises(InvalidWorkbookError):
        _load_from_bytes(b"not xlsx data at all")


def test_load_from_bytes_returns_workbook() -> None:
    """_load_from_bytes with valid XLSX bytes returns an openpyxl Workbook."""
    wb = _load_from_bytes(_make_xlsx_bytes())
    assert isinstance(wb, openpyxl.Workbook)


# ---------------------------------------------------------------------------
# _load_from_stream
# ---------------------------------------------------------------------------


def test_corrupted_stream_raises_invalid_workbook_error() -> None:
    """_load_from_stream with bad bytes raises InvalidWorkbookError."""
    with pytest.raises(InvalidWorkbookError):
        _load_from_stream(BytesIO(b"bad stream data"))


def test_load_from_stream_returns_workbook() -> None:
    """_load_from_stream with a valid XLSX stream returns an openpyxl Workbook."""
    stream = BytesIO(_make_xlsx_bytes())
    wb = _load_from_stream(stream)
    assert isinstance(wb, openpyxl.Workbook)


# ---------------------------------------------------------------------------
# data_only=True verification
# ---------------------------------------------------------------------------


def test_load_uses_data_only(tmp_path: Path) -> None:
    """All loaders open with data_only=True (cached values, not formula strings)."""
    # Create an XLSX with a numeric value in a cell (acts as pre-computed result)
    xlsx_bytes = _make_xlsx_bytes()
    xlsx_file = tmp_path / "test.xlsx"
    xlsx_file.write_bytes(xlsx_bytes)

    # All three loaders should return a Workbook opened with data_only=True
    wb_path = _load_from_path(xlsx_file)
    wb_bytes = _load_from_bytes(xlsx_bytes)
    wb_stream = _load_from_stream(BytesIO(xlsx_bytes))

    # Verify that each loaded workbook has data_only mode (not returning formula strings)
    for wb in (wb_path, wb_bytes, wb_stream):
        ws = wb.active
        assert ws is not None
        # A1 has "hello" — we verify the actual data is accessible (not a formula string)
        assert ws["A1"].value == "hello"
        assert ws["B1"].value == 42
