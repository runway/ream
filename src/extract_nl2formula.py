"""
Extract QA pairs from NL2Formula dataset.
Converts 2D table arrays to XLSX files, evaluates simple formulas to get ground truth values.
"""

import json
import os
import re
import tempfile
from typing import Any

import openpyxl


def evaluate_simple_formula(formula: str, table: list[list[str]]) -> str | None:
    """Evaluate a simple NL2Formula formula against a 2D table.

    Handles common patterns:
    - FILTER(col, condition) → filtered values
    - MIN/MAX/SUM/AVERAGE over filtered results
    - ROWS(FILTER(...)) → count
    - UNIQUE(CHOOSECOLS(FILTER(...), n)) → unique values from column n
    - Simple FILTER returning values

    Returns the string result or None if formula is too complex.
    """
    # Table is: row 0 = ["0", "A", "B", ...], row 1 = ["1", header1, header2, ...]
    # row 2+ = ["2", val1, val2, ...]
    if len(table) < 3:
        return None

    ncols = len(table[0]) - 1  # exclude row index column
    nrows = len(table) - 1     # exclude header row (row 0 is column letters)

    def col_idx(letter: str) -> int:
        """Convert column letter to 0-based index in data."""
        letter = letter.upper()
        idx = 0
        for c in letter:
            idx = idx * 26 + (ord(c) - ord('A') + 1)
        return idx  # 1-based matches table column position

    def get_col_data(col_letter: str) -> list[str]:
        """Get all data values for a column (skipping header rows)."""
        ci = col_idx(col_letter)
        if ci >= len(table[0]):
            return []
        return [table[r][ci] for r in range(2, len(table)) if ci < len(table[r])]

    def parse_number(s: str) -> float | None:
        s = s.strip().replace(",", "")
        try:
            return float(s)
        except (ValueError, TypeError):
            return None

    formula = formula.strip()

    # Pattern: FILTER(col_range, condition) - simple single-column filter
    # E.g., FILTER(D2:D18, H2:H18="seattle center coliseum 11,497")
    m = re.match(r'^FILTER\(([A-Z])(\d+):([A-Z])(\d+)\s*,\s*([A-Z])(\d+):([A-Z])(\d+)\s*=\s*"?([^"]*?)"?\s*\)$', formula)
    if m:
        result_col = col_idx(m.group(1))
        filter_col = col_idx(m.group(5))
        filter_val = m.group(9).strip().lower()
        results = []
        for r in range(2, len(table)):
            if filter_col < len(table[r]) and result_col < len(table[r]):
                if table[r][filter_col].strip().lower() == filter_val:
                    results.append(table[r][result_col])
        if results:
            return results[0] if len(results) == 1 else ", ".join(results)
        return None

    # Pattern: UNIQUE(CHOOSECOLS(FILTER(range, condition), n))
    m = re.match(r'^UNIQUE\(CHOOSECOLS\(FILTER\(([A-Z])(\d+):([A-Z])(\d+)\s*,\s*([A-Z])(\d+):([A-Z])(\d+)\s*=\s*"?([^"]*?)"?\s*\)\s*,\s*(\d+)\)\)', formula)
    if m:
        filter_col = col_idx(m.group(5))
        filter_val = m.group(9).strip().lower()
        choose_col_offset = int(m.group(10))
        start_col = col_idx(m.group(1))
        target_col = start_col + choose_col_offset - 1
        results = []
        for r in range(2, len(table)):
            if filter_col < len(table[r]) and target_col < len(table[r]):
                if table[r][filter_col].strip().lower() == filter_val:
                    val = table[r][target_col]
                    if val not in results:
                        results.append(val)
        if results:
            return results[0] if len(results) == 1 else ", ".join(results)
        return None

    # Pattern: MIN/MAX/SUM(FILTER(col, condition))
    m = re.match(r'^(MIN|MAX|SUM|AVERAGE)\(FILTER\(([A-Z])(\d+):([A-Z])(\d+)\s*,\s*(.+)\)\)$', formula)
    if m:
        agg = m.group(1)
        result_col = col_idx(m.group(2))
        cond_str = m.group(6).strip()

        # Parse simple equality conditions: col="val" or (col=val)*(col2=val2)
        conditions = []
        for part in re.findall(r'([A-Z])(\d+):([A-Z])(\d+)\s*=\s*"?([^")*]*?)"?(?:\)|$|\*)', cond_str):
            conditions.append((col_idx(part[0]), part[4].strip().lower()))

        if not conditions:
            # Try simpler pattern: col=val
            for part in re.findall(r'([A-Z])(\d+):([A-Z])(\d+)\s*=\s*(\d+\.?\d*)', cond_str):
                conditions.append((col_idx(part[0]), part[4].strip()))

        if conditions:
            results = []
            for r in range(2, len(table)):
                match = True
                for cc, cv in conditions:
                    if cc >= len(table[r]) or table[r][cc].strip().lower() != cv:
                        match = False
                        break
                if match and result_col < len(table[r]):
                    n = parse_number(table[r][result_col])
                    if n is not None:
                        results.append(n)

            if results:
                if agg == "MIN":
                    return str(min(results))
                elif agg == "MAX":
                    return str(max(results))
                elif agg == "SUM":
                    return str(sum(results))
                elif agg == "AVERAGE":
                    return str(sum(results) / len(results))

    # Pattern: ROWS(FILTER(...)) or ROWS(UNIQUE(FILTER(...))) - counting
    m = re.match(r'^ROWS\((UNIQUE\()?FILTER\(([A-Z])(\d+):([A-Z])(\d+)\s*,\s*([A-Z])(\d+):([A-Z])(\d+)\s*=\s*"?([^"]*?)"?\s*\)\)?\)$', formula)
    if m:
        filter_col = col_idx(m.group(6))
        filter_val = m.group(10).strip().lower()
        is_unique = m.group(1) is not None
        results = []
        for r in range(2, len(table)):
            if filter_col < len(table[r]):
                if table[r][filter_col].strip().lower() == filter_val:
                    results.append(r)
        count = len(set(results)) if is_unique else len(results)
        return str(count)

    return None


def table_to_xlsx(table: list[list[str]], output_path: str):
    """Write a NL2Formula 2D table to an XLSX file."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"

    # Table format: row 0 = ["0", "A", "B", ...] (column letters)
    # row 1 = ["1", header1, header2, ...] (headers)
    # row 2+ = ["2", val1, val2, ...] (data)
    for r_idx, row in enumerate(table):
        if r_idx == 0:
            continue  # Skip the column-letter row
        for c_idx, val in enumerate(row):
            if c_idx == 0:
                continue  # Skip the row-number column
            # Try to write as number if possible
            try:
                ws.cell(row=r_idx, column=c_idx, value=float(val))
            except (ValueError, TypeError):
                ws.cell(row=r_idx, column=c_idx, value=str(val))

    wb.save(output_path)
    wb.close()


def extract_nl2formula_questions(corpus_dir: str, max_questions: int = 300) -> list[dict]:
    """Extract QA pairs from NL2Formula with evaluated ground truth."""
    test_path = os.path.join(corpus_dir, "dataset/test.json")
    with open(test_path) as f:
        data = json.load(f)

    xlsx_dir = os.path.join(corpus_dir, "xlsx_cache")
    os.makedirs(xlsx_dir, exist_ok=True)

    questions = []
    for item in data:
        if len(questions) >= max_questions:
            break

        table = item["Table"]
        table_name = item["TableName"]
        xlsx_path = os.path.join(xlsx_dir, f"{table_name}.xlsx")

        # Write XLSX if not cached
        if not os.path.exists(xlsx_path):
            table_to_xlsx(table, xlsx_path)

        for formula_item in item["t5Formulas"]:
            if len(questions) >= max_questions:
                break

            question = formula_item["Question"]
            formula2 = formula_item.get("Formula2", formula_item["Formula"])
            level = formula_item.get("Level", "unknown")

            # Try to evaluate the formula to get ground truth
            answer = evaluate_simple_formula(formula2, table)
            if answer is None:
                continue  # Skip formulas we can't evaluate

            # Clean up answer
            answer = answer.strip()
            if not answer:
                continue

            questions.append({
                "workbook": f"nl2f_{table_name}",
                "workbook_path": os.path.abspath(xlsx_path),
                "question": question,
                "answer_resolved": answer,
                "formula": formula2,
                "question_type": f"nl2formula_{level}",
                "difficulty": {"easy": "Easy", "medium": "Medium", "hard": "Hard", "Calculation": "Medium"}.get(level, "Medium"),
                "sheet": "Sheet1",
            })

    return questions


if __name__ == "__main__":
    corpus_dir = "corpus/nl2formula"
    questions = extract_nl2formula_questions(corpus_dir, max_questions=500)
    print(f"Extracted {len(questions)} evaluable NL2Formula questions")

    types = {}
    diffs = {}
    for q in questions:
        types[q["question_type"]] = types.get(q["question_type"], 0) + 1
        diffs[q["difficulty"]] = diffs.get(q["difficulty"], 0) + 1
    print(f"Types: {json.dumps(types, indent=2)}")
    print(f"Difficulties: {json.dumps(diffs, indent=2)}")

    print("\nSamples:")
    for q in questions[:8]:
        print(f"  [{q['difficulty']}] Q: {q['question'][:70]}")
        print(f"    A: {q['answer_resolved'][:60]}")
        print(f"    F: {q['formula'][:60]}")
        print()

    with open("data/nl2formula_questions.json", "w") as f:
        json.dump(questions, f, indent=2, default=str)
    print(f"Saved to data/nl2formula_questions.json")
