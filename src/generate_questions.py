"""
Generate a large corpus of QA tasks from FRTR workbooks and SpreadsheetBench data.
Targets 1000+ questions with verified ground truth.

Question types:
  - Cell lookup: "What is the value in cell [Sheet]![Cell]?"
  - Row lookup: "What is the value in column [Col] of row [N] in sheet [Sheet]?"
  - Aggregation: "What is the sum/max/min/avg of column [Col] in sheet [Sheet]?"
  - Count: "How many non-empty rows are in sheet [Sheet]?"
  - Filter: "What is the [Col] value where [OtherCol] = [Value]?"
  - Cross-sheet: "What is the value in [Sheet2] that corresponds to [Value] in [Sheet1]?"
  - Header: "What are the column headers in sheet [Sheet]?"
  - Schema: "How many sheets does this workbook contain?"
"""

import os
import re
import json
import random
from datetime import datetime, date
from typing import Any

import openpyxl
from openpyxl.utils import get_column_letter


def _fmt(val: Any) -> str:
    """Format a value for ground truth comparison."""
    if val is None:
        return ""
    if isinstance(val, bool):
        return "TRUE" if val else "FALSE"
    if isinstance(val, float):
        if val == int(val) and abs(val) < 1e15:
            return str(int(val))
        return f"{val:.6f}".rstrip("0").rstrip(".")
    if isinstance(val, datetime):
        return val.strftime("%Y-%m-%d")
    if isinstance(val, date):
        return val.strftime("%Y-%m-%d")
    return str(val)


def _is_data_sheet(ws, sheet_name: str) -> bool:
    """Check if a sheet contains tabular data (not metadata/questions)."""
    skip = {"Questions", "Readme", "README", "Metadata", "Info"}
    if sheet_name in skip:
        return False
    if ws.max_row is None or ws.max_row < 3:
        return False
    if ws.max_column is None or ws.max_column < 2:
        return False
    return True


def _get_header_row(ws) -> tuple[int, dict[int, str]]:
    """Find the header row and return (row_idx, {col_idx: header_name})."""
    for row_idx in range(1, min(5, (ws.max_row or 1) + 1)):
        headers = {}
        for col_idx in range(1, (ws.max_column or 0) + 1):
            val = ws.cell(row=row_idx, column=col_idx).value
            if val is not None and isinstance(val, str) and len(val) < 50:
                headers[col_idx] = val
        if len(headers) >= 2:
            return row_idx, headers
    return 0, {}


def _get_numeric_columns(ws, header_row: int, headers: dict[int, str], max_rows: int = 200) -> dict[int, list[float]]:
    """Find columns that contain numeric data and collect their values."""
    numeric_cols = {}
    for col_idx in headers:
        vals = []
        for row_idx in range(header_row + 1, min((ws.max_row or 0) + 1, header_row + max_rows + 1)):
            v = ws.cell(row=row_idx, column=col_idx).value
            if isinstance(v, (int, float)) and not isinstance(v, bool):
                vals.append(float(v))
        if len(vals) >= 3:
            numeric_cols[col_idx] = vals
    return numeric_cols


def _get_text_columns(ws, header_row: int, headers: dict[int, str], max_rows: int = 200) -> dict[int, list[str]]:
    """Find columns with text data."""
    text_cols = {}
    for col_idx in headers:
        vals = []
        for row_idx in range(header_row + 1, min((ws.max_row or 0) + 1, header_row + max_rows + 1)):
            v = ws.cell(row=row_idx, column=col_idx).value
            if isinstance(v, str) and len(v) < 100:
                vals.append(v)
        if len(vals) >= 3:
            text_cols[col_idx] = vals
    return text_cols


def generate_cell_lookup_questions(ws, sheet_name: str, header_row: int, headers: dict, max_rows: int = 200) -> list[dict]:
    """Generate 'What is the value in cell X?' questions."""
    questions = []
    data_rows = list(range(header_row + 1, min((ws.max_row or 0) + 1, header_row + max_rows + 1)))
    if not data_rows:
        return []

    # Sample up to 5 random cells per sheet
    sample_rows = random.sample(data_rows, min(5, len(data_rows)))
    sample_cols = list(headers.keys())

    for row_idx in sample_rows:
        col_idx = random.choice(sample_cols)
        val = ws.cell(row=row_idx, column=col_idx).value
        if val is None:
            continue
        col_letter = get_column_letter(col_idx)
        header_name = headers.get(col_idx, col_letter)

        questions.append({
            "question": f"What is the value in the '{header_name}' column of row {row_idx} in sheet '{sheet_name}'?",
            "answer_resolved": _fmt(val),
            "question_type": "cell_lookup",
            "difficulty": "Easy",
            "sheet": sheet_name,
        })

    return questions


def generate_aggregation_questions(ws, sheet_name: str, header_row: int, headers: dict, max_rows: int = 200) -> list[dict]:
    """Generate aggregation questions (sum, max, min, average, count)."""
    questions = []
    numeric_cols = _get_numeric_columns(ws, header_row, headers, max_rows)

    for col_idx, vals in numeric_cols.items():
        header_name = headers.get(col_idx, get_column_letter(col_idx))

        # Sum
        total = sum(vals)
        questions.append({
            "question": f"What is the sum of all values in the '{header_name}' column in sheet '{sheet_name}'? Consider only the first {max_rows} data rows.",
            "answer_resolved": _fmt(total),
            "question_type": "aggregation_sum",
            "difficulty": "Medium",
            "sheet": sheet_name,
        })

        # Max
        max_val = max(vals)
        questions.append({
            "question": f"What is the maximum value in the '{header_name}' column in sheet '{sheet_name}'? Answer based only on the data shown.",
            "answer_resolved": _fmt(max_val),
            "question_type": "aggregation_max",
            "difficulty": "Easy",
            "sheet": sheet_name,
        })

        # Min
        min_val = min(vals)
        questions.append({
            "question": f"What is the minimum value in the '{header_name}' column in sheet '{sheet_name}'? Answer based only on the data shown.",
            "answer_resolved": _fmt(min_val),
            "question_type": "aggregation_min",
            "difficulty": "Easy",
            "sheet": sheet_name,
        })

        # Average
        avg_val = total / len(vals)
        questions.append({
            "question": f"What is the average of the '{header_name}' column in sheet '{sheet_name}'? Consider only the first {max_rows} data rows.",
            "answer_resolved": _fmt(avg_val),
            "question_type": "aggregation_avg",
            "difficulty": "Medium",
            "sheet": sheet_name,
        })

        # Count
        questions.append({
            "question": f"How many numeric values are in the '{header_name}' column in sheet '{sheet_name}' (first {max_rows} data rows)?",
            "answer_resolved": str(len(vals)),
            "question_type": "count",
            "difficulty": "Easy",
            "sheet": sheet_name,
        })

    return questions


def generate_filter_questions(ws, sheet_name: str, header_row: int, headers: dict, max_rows: int = 200) -> list[dict]:
    """Generate filter/lookup questions: 'What is X where Y = Z?'"""
    questions = []
    text_cols = _get_text_columns(ws, header_row, headers, max_rows)
    numeric_cols = _get_numeric_columns(ws, header_row, headers, max_rows)

    if not text_cols or not numeric_cols:
        return []

    # Pick a text column as the filter and a numeric column as the target
    text_col_idx = random.choice(list(text_cols.keys()))
    text_header = headers.get(text_col_idx, get_column_letter(text_col_idx))
    unique_vals = list(set(text_cols[text_col_idx]))

    if len(unique_vals) < 2:
        return []

    for num_col_idx in list(numeric_cols.keys())[:2]:
        num_header = headers.get(num_col_idx, get_column_letter(num_col_idx))

        # Pick up to 3 filter values
        filter_vals = random.sample(unique_vals, min(3, len(unique_vals)))

        for filter_val in filter_vals:
            # Find matching rows and compute sum
            matching_sum = 0
            match_count = 0
            for row_idx in range(header_row + 1, min((ws.max_row or 0) + 1, header_row + max_rows + 1)):
                text_val = ws.cell(row=row_idx, column=text_col_idx).value
                num_val = ws.cell(row=row_idx, column=num_col_idx).value
                if str(text_val) == filter_val and isinstance(num_val, (int, float)):
                    matching_sum += float(num_val)
                    match_count += 1

            if match_count > 0:
                questions.append({
                    "question": f"What is the sum of '{num_header}' where '{text_header}' equals '{filter_val}' in sheet '{sheet_name}'?",
                    "answer_resolved": _fmt(matching_sum),
                    "question_type": "filter_sum",
                    "difficulty": "Medium",
                    "sheet": sheet_name,
                })

                questions.append({
                    "question": f"How many rows have '{text_header}' equal to '{filter_val}' in sheet '{sheet_name}'?",
                    "answer_resolved": str(match_count),
                    "question_type": "filter_count",
                    "difficulty": "Medium",
                    "sheet": sheet_name,
                })

    return questions


def generate_schema_questions(wb, filepath: str) -> list[dict]:
    """Generate workbook-level schema questions."""
    questions = []
    data_sheets = [s for s in wb.sheetnames if s not in ("Questions", "Readme", "README")]

    questions.append({
        "question": "How many sheets does this workbook contain (excluding any Questions or Readme sheets)?",
        "answer_resolved": str(len(data_sheets)),
        "question_type": "schema_count",
        "difficulty": "Easy",
        "sheet": "_workbook",
    })

    questions.append({
        "question": "What are the names of the data sheets in this workbook (excluding Questions and Readme)?",
        "answer_resolved": ", ".join(data_sheets),
        "question_type": "schema_names",
        "difficulty": "Easy",
        "sheet": "_workbook",
    })

    for sheet_name in data_sheets:
        ws = wb[sheet_name]
        header_row, headers = _get_header_row(ws)
        if headers:
            header_names = [headers[k] for k in sorted(headers.keys())]
            questions.append({
                "question": f"What are the column headers in sheet '{sheet_name}'?",
                "answer_resolved": ", ".join(header_names),
                "question_type": "schema_headers",
                "difficulty": "Easy",
                "sheet": sheet_name,
            })

    return questions


def generate_cross_sheet_questions(wb, data_sheets: list[str], max_rows: int = 200) -> list[dict]:
    """Generate questions requiring cross-sheet reasoning."""
    questions = []

    if len(data_sheets) < 2:
        return []

    # Find sheets with shared column headers (potential join keys)
    sheet_headers = {}
    for sheet_name in data_sheets:
        ws = wb[sheet_name]
        _, headers = _get_header_row(ws)
        sheet_headers[sheet_name] = {v for v in headers.values()}

    for i, s1 in enumerate(data_sheets):
        for s2 in data_sheets[i + 1:]:
            shared = sheet_headers.get(s1, set()) & sheet_headers.get(s2, set())
            if not shared:
                continue

            questions.append({
                "question": f"Which column headers are shared between sheets '{s1}' and '{s2}'?",
                "answer_resolved": ", ".join(sorted(shared)),
                "question_type": "cross_sheet_schema",
                "difficulty": "Medium",
                "sheet": f"{s1}+{s2}",
            })

    return questions


def generate_from_frtr(corpus_dir: str, max_rows: int = 200, seed: int = 42) -> list[dict]:
    """Generate a large question corpus from FRTR workbooks."""
    random.seed(seed)
    all_questions = []

    xlsx_files = sorted(f for f in os.listdir(corpus_dir) if f.endswith(".xlsx") and f.startswith("frtr_"))

    for fname in xlsx_files:
        fpath = os.path.join(corpus_dir, fname)
        try:
            wb = openpyxl.load_workbook(fpath, data_only=True)
        except Exception as e:
            print(f"  WARN: Could not load {fname}: {e}")
            continue

        data_sheets = [s for s in wb.sheetnames if _is_data_sheet(wb[s], s)]

        # Filter to sheets that fit within max_rows (no truncation)
        fitting_sheets = []
        for s in data_sheets:
            ws = wb[s]
            if (ws.max_row or 0) <= max_rows:
                fitting_sheets.append(s)

        # Schema questions (use all data_sheets for names, but only fitting for content)
        schema_qs = generate_schema_questions(wb, fpath)
        for q in schema_qs:
            q["workbook"] = fname
            q["workbook_path"] = os.path.abspath(fpath)
        all_questions.extend(schema_qs)

        # Per-sheet questions — only on sheets that fit fully
        for sheet_name in fitting_sheets:
            ws = wb[sheet_name]
            header_row, headers = _get_header_row(ws)
            if not headers:
                continue

            # Cell lookups
            cell_qs = generate_cell_lookup_questions(ws, sheet_name, header_row, headers, max_rows)
            for q in cell_qs:
                q["workbook"] = fname
                q["workbook_path"] = os.path.abspath(fpath)
            all_questions.extend(cell_qs)

            # Aggregations
            agg_qs = generate_aggregation_questions(ws, sheet_name, header_row, headers, max_rows)
            for q in agg_qs:
                q["workbook"] = fname
                q["workbook_path"] = os.path.abspath(fpath)
            all_questions.extend(agg_qs)

            # Filter questions
            filter_qs = generate_filter_questions(ws, sheet_name, header_row, headers, max_rows)
            for q in filter_qs:
                q["workbook"] = fname
                q["workbook_path"] = os.path.abspath(fpath)
            all_questions.extend(filter_qs)

        # Cross-sheet questions — only between sheets that fit
        cross_qs = generate_cross_sheet_questions(wb, fitting_sheets, max_rows)
        for q in cross_qs:
            q["workbook"] = fname
            q["workbook_path"] = os.path.abspath(fpath)
        all_questions.extend(cross_qs)

        wb.close()

    return all_questions


def generate_from_spreadsheetbench(data_dir: str, max_tasks: int = 200, seed: int = 42) -> list[dict]:
    """Generate QA tasks from SpreadsheetBench input/answer pairs.

    Each task asks about the expected output value at the answer position.
    """
    random.seed(seed)
    questions = []

    dataset_path = os.path.join(data_dir, "dataset.json")
    if not os.path.exists(dataset_path):
        print(f"  WARN: {dataset_path} not found")
        return []

    with open(dataset_path) as f:
        dataset = json.load(f)

    for item in dataset[:max_tasks]:
        task_id = item["id"]
        instruction = item["instruction"]
        answer_pos = item.get("answer_position", "")
        spreadsheet_dir = os.path.join(data_dir, item["spreadsheet_path"])

        if not os.path.isdir(spreadsheet_dir):
            continue

        # Find input and answer files
        input_files = sorted(f for f in os.listdir(spreadsheet_dir) if f.endswith("_input.xlsx"))
        answer_files = sorted(f for f in os.listdir(spreadsheet_dir) if f.endswith("_answer.xlsx"))

        if not input_files or not answer_files:
            continue

        input_path = os.path.join(spreadsheet_dir, input_files[0])
        answer_path = os.path.join(spreadsheet_dir, answer_files[0])

        try:
            answer_wb = openpyxl.load_workbook(answer_path, data_only=True)
            answer_ws = answer_wb.active

            # Parse the answer position to get the expected value
            if answer_pos and ":" in answer_pos:
                # Range like "H3:H5" - get the first cell
                first_cell = answer_pos.split(":")[0]
                val = answer_ws[first_cell].value
            elif answer_pos:
                val = answer_ws[answer_pos].value
            else:
                val = None

            answer_wb.close()

            if val is None:
                continue

            questions.append({
                "workbook": f"sb_{task_id}",
                "workbook_path": os.path.abspath(input_path),
                "question": f"Given this spreadsheet, {instruction.strip()} What value should appear at cell {answer_pos.split(':')[0] if ':' in answer_pos else answer_pos}?",
                "answer_resolved": _fmt(val),
                "question_type": "spreadsheetbench_task",
                "difficulty": "Hard",
                "sheet": "Sheet1",
            })

        except Exception as e:
            continue

    return questions


if __name__ == "__main__":
    import sys

    frtr_dir = sys.argv[1] if len(sys.argv) > 1 else "corpus/frtr"
    sb_dir = sys.argv[2] if len(sys.argv) > 2 else "corpus/spreadsheetbench/data/sample_data_200"

    max_rows = int(sys.argv[3]) if len(sys.argv) > 3 else 1000

    print(f"Generating FRTR questions (max_rows={max_rows})...")
    frtr_qs = generate_from_frtr(frtr_dir, max_rows=max_rows)
    print(f"  Generated {len(frtr_qs)} FRTR questions")

    print("Generating SpreadsheetBench questions...")
    sb_qs = generate_from_spreadsheetbench(sb_dir)
    print(f"  Generated {len(sb_qs)} SpreadsheetBench questions")

    all_qs = frtr_qs + sb_qs
    print(f"\nTotal: {len(all_qs)} questions")

    # Stats
    types = {}
    diffs = {}
    for q in all_qs:
        types[q["question_type"]] = types.get(q["question_type"], 0) + 1
        diffs[q["difficulty"]] = diffs.get(q["difficulty"], 0) + 1

    print(f"By type: {json.dumps(types, indent=2)}")
    print(f"By difficulty: {json.dumps(diffs, indent=2)}")
    print(f"Unique workbooks: {len(set(q['workbook'] for q in all_qs))}")

    # Save
    os.makedirs("data", exist_ok=True)
    with open("data/questions_large.json", "w") as f:
        json.dump(all_qs, f, indent=2, default=str)
    print(f"\nSaved to data/questions_large.json")
