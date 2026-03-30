"""
Convert XLSX workbooks to various text representations for LLM consumption.
Supports: Ream, SCF (legacy), CSV, Markdown, JSON and others
"""

import json
import re
from datetime import datetime, date
from typing import Any

import openpyxl
from openpyxl.utils import get_column_letter


def _cell_value_str(val: Any) -> str:
    """Convert a cell value to a string representation."""
    if val is None:
        return ""
    if isinstance(val, bool):
        return "TRUE" if val else "FALSE"
    if isinstance(val, datetime):
        return val.strftime("%Y-%m-%d")
    if isinstance(val, date):
        return val.strftime("%Y-%m-%d")
    if isinstance(val, float):
        # Remove trailing zeros
        if val == int(val) and abs(val) < 1e15:
            return str(int(val))
        return f"{val:.10g}"
    return str(val)


def _needs_scf_quoting(s: str) -> bool:
    """Check if a string needs quoting in SCF format."""
    if not s:
        return True  # empty string
    if s.startswith("=") or s.startswith("[") or s.startswith('"'):
        return True
    if "|" in s or "\n" in s or "\r" in s:
        return True
    if s != s.strip():
        return True  # leading/trailing whitespace
    if s in ("TRUE", "FALSE", "true", "false"):
        return True
    # Check if it looks like a number
    try:
        float(s)
        return True
    except ValueError:
        pass
    # Check if it looks like an error
    if s in ("#N/A", "#REF!", "#VALUE!", "#DIV/0!", "#NULL!", "#NAME?", "#NUM!"):
        return True
    # Check if it matches addressed-entry pattern
    if re.match(r"^C[1-9][0-9]*(?::C[1-9][0-9]*)?=", s):
        return True
    return False


def _scf_quote(s: str) -> str:
    """Quote a string for SCF format."""
    escaped = s.replace("\\", "\\\\").replace('"', '""').replace("\n", "\\n").replace("\r", "\\r").replace("\t", "\\t")
    return f'"{escaped}"'


def _scf_scalar(val: Any) -> str:
    """Convert a cell value to an SCF scalar."""
    if val is None:
        return ""
    if isinstance(val, bool):
        return "TRUE" if val else "FALSE"
    if isinstance(val, (int, float)):
        if isinstance(val, float):
            if val == int(val) and abs(val) < 1e15:
                return str(int(val))
            return f"{val:.10g}"
        return str(val)
    if isinstance(val, datetime):
        s = val.strftime("%Y-%m-%d")
        return s
    if isinstance(val, date):
        s = val.strftime("%Y-%m-%d")
        return s
    s = str(val)
    if _needs_scf_quoting(s):
        return _scf_quote(s)
    return s


def xlsx_to_scf(filepath: str, max_rows_per_sheet: int = 500, addressed: bool = False) -> str:
    """Convert an XLSX workbook to SCF format.

    Args:
        filepath: Path to the XLSX workbook.
        max_rows_per_sheet: Truncate sheets to this many data rows.
        addressed: If True, every cell gets a C<N>= prefix (eliminates
                   positional counting). If False, only sparse gaps get prefixes.
    """
    wb = openpyxl.load_workbook(filepath, data_only=True)
    lines = ["#!SCF 5"]

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        # Quote sheet name if it contains special chars
        safe_name = sheet_name
        if " " in sheet_name or "|" in sheet_name or '"' in sheet_name or "&" in sheet_name:
            safe_name = _scf_quote(sheet_name)
        lines.append(f"#!SHEET {safe_name}")

        # Detect header row (first row with data)
        header_row = None
        for row_idx in range(1, min(ws.max_row or 1, max_rows_per_sheet) + 1):
            has_data = False
            for col_idx in range(1, (ws.max_column or 1) + 1):
                if ws.cell(row=row_idx, column=col_idx).value is not None:
                    has_data = True
                    break
            if has_data:
                header_row = row_idx
                break

        if header_row:
            lines.append(f"#!HEADERS R{header_row}")

        row_count = 0
        for row_idx in range(1, (ws.max_row or 0) + 1):
            if row_count >= max_rows_per_sheet:
                lines.append(f"# ... showing first {max_rows_per_sheet} data rows")
                break

            # Collect non-empty cells in this row
            cells = {}
            for col_idx in range(1, (ws.max_column or 0) + 1):
                val = ws.cell(row=row_idx, column=col_idx).value
                if val is not None:
                    cells[col_idx] = val

            if not cells:
                continue

            row_count += 1
            # Build row record
            entries = []
            if addressed:
                # Every cell gets an explicit C<N>= prefix
                for col_idx in sorted(cells.keys()):
                    scalar = _scf_scalar(cells[col_idx])
                    entries.append(f"C{col_idx}={scalar}")
            else:
                # SCF canonical: only sparse gaps get prefixes
                cursor = 1
                for col_idx in sorted(cells.keys()):
                    scalar = _scf_scalar(cells[col_idx])
                    if col_idx == cursor:
                        entries.append(scalar)
                    else:
                        entries.append(f"C{col_idx}={scalar}")
                    cursor = col_idx + 1

            line = f"R{row_idx} | " + " | ".join(entries) + " |"
            lines.append(line)

    return "\n".join(lines)


def xlsx_to_scf_addressed(filepath: str, max_rows_per_sheet: int = 500) -> str:
    """SCF with always-on column addressing (every cell prefixed with C<N>=)."""
    return xlsx_to_scf(filepath, max_rows_per_sheet=max_rows_per_sheet, addressed=True)


def xlsx_to_csv(filepath: str, max_rows_per_sheet: int = 500) -> str:
    """Convert an XLSX workbook to CSV-style text (one section per sheet)."""
    wb = openpyxl.load_workbook(filepath, data_only=True)
    sections = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows_out = [f"=== Sheet: {sheet_name} ==="]
        row_count = 0

        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=True):
            if row_count >= max_rows_per_sheet:
                rows_out.append(f"... showing first {max_rows_per_sheet} data rows")
                break
            vals = [_cell_value_str(v) for v in row]
            # Skip completely empty rows
            if all(v == "" for v in vals):
                continue
            rows_out.append(",".join(vals))
            row_count += 1

        sections.append("\n".join(rows_out))

    return "\n\n".join(sections)


def xlsx_to_markdown(filepath: str, max_rows_per_sheet: int = 500) -> str:
    """Convert an XLSX workbook to Markdown table format."""
    wb = openpyxl.load_workbook(filepath, data_only=True)
    sections = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        section_lines = [f"## Sheet: {sheet_name}\n"]

        all_rows = []
        row_count = 0
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=True):
            if row_count >= max_rows_per_sheet:
                break
            vals = [_cell_value_str(v) for v in row]
            if all(v == "" for v in vals):
                continue
            all_rows.append(vals)
            row_count += 1

        if not all_rows:
            sections.append(f"## Sheet: {sheet_name}\n\n(empty)")
            continue

        # Use first row as header
        max_cols = max(len(r) for r in all_rows)
        for i, r in enumerate(all_rows):
            while len(r) < max_cols:
                all_rows[i] = list(r) + [""]

        header = all_rows[0]
        section_lines.append("| " + " | ".join(str(h) for h in header) + " |")
        section_lines.append("| " + " | ".join("---" for _ in header) + " |")

        for r in all_rows[1:]:
            section_lines.append("| " + " | ".join(str(v) for v in r) + " |")

        if row_count >= max_rows_per_sheet and ws.max_row > max_rows_per_sheet:
            section_lines.append(f"\n*... showing first {max_rows_per_sheet} data rows*")

        sections.append("\n".join(section_lines))

    return "\n\n".join(sections)


def xlsx_to_json(filepath: str, max_rows_per_sheet: int = 500) -> str:
    """Convert an XLSX workbook to structured JSON."""
    wb = openpyxl.load_workbook(filepath, data_only=True)
    workbook_data = {"sheets": {}}

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        sheet_data = {"rows": [], "dimensions": {"max_row": ws.max_row, "max_col": ws.max_column}}

        row_count = 0
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=False):
            if row_count >= max_rows_per_sheet:
                sheet_data["truncated"] = True
                sheet_data["rows_shown"] = max_rows_per_sheet
                break

            row_data = {}
            for cell in row:
                if cell.value is not None:
                    col_letter = get_column_letter(cell.column)
                    row_data[col_letter] = _cell_value_str(cell.value)

            if row_data:
                row_data["_row"] = row[0].row
                sheet_data["rows"].append(row_data)
                row_count += 1

        workbook_data["sheets"][sheet_name] = sheet_data

    return json.dumps(workbook_data, indent=2, ensure_ascii=False)


def _classify_dtype(val: Any) -> str:
    """Classify a cell value into a data type category for SheetCompressor."""
    if val is None:
        return ""
    if isinstance(val, bool):
        return "Boolean"
    if isinstance(val, datetime):
        return "DateData"
    if isinstance(val, date):
        return "DateData"
    if isinstance(val, int):
        return "IntNum"
    if isinstance(val, float):
        if val == int(val) and abs(val) < 1e15:
            return "IntNum"
        return "FloatNum"
    s = str(val)
    if not s:
        return ""
    # Check for percentage pattern
    if s.endswith("%") or (len(s) > 1 and s[-1] == "%" and s[:-1].replace(".", "").replace("-", "").isdigit()):
        return "PctNum"
    # Check for currency
    if s.startswith("$") or s.startswith("€") or s.startswith("£"):
        return "Currency"
    # Check for email
    if "@" in s and "." in s:
        return "EmailData"
    # Check for date-like strings
    if re.match(r"^\d{4}[-/]\d{2}[-/]\d{2}", s):
        return "DateData"
    return "Text"


def _merge_ranges(addresses: list[str]) -> str:
    """Merge a list of cell addresses into compact range notation.

    E.g., ['A1', 'A2', 'A3', 'B1', 'B2', 'B3'] -> 'A1:B3'
    """
    if not addresses:
        return ""
    if len(addresses) == 1:
        return addresses[0]

    # Parse addresses into (col_letter, row_num) tuples
    parsed = []
    for addr in addresses:
        m = re.match(r"^([A-Z]+)(\d+)$", addr)
        if m:
            parsed.append((m.group(1), int(m.group(2))))

    if not parsed:
        return addresses[0]

    # Try to find rectangular regions
    cols = sorted(set(p[0] for p in parsed))
    rows = sorted(set(p[1] for p in parsed))

    # Check if all cells form a contiguous rectangle
    addr_set = set((c, r) for c, r in parsed)
    is_rect = True
    for c in cols:
        for r in rows:
            if (c, r) not in addr_set:
                is_rect = False
                break
        if not is_rect:
            break

    if is_rect and len(cols) > 0 and len(rows) > 0:
        # Check if rows and cols are contiguous
        min_col, max_col = cols[0], cols[-1]
        min_row, max_row = rows[0], rows[-1]
        return f"{min_col}{min_row}:{max_col}{max_row}"

    # Fall back to listing contiguous column runs
    segments = []
    parsed_sorted = sorted(parsed, key=lambda x: (x[0], x[1]))

    current_col = None
    current_start = None
    current_end = None

    for col, row in parsed_sorted:
        if col == current_col and row == current_end + 1:
            current_end = row
        else:
            if current_col is not None:
                if current_start == current_end:
                    segments.append(f"{current_col}{current_start}")
                else:
                    segments.append(f"{current_col}{current_start}:{current_col}{current_end}")
            current_col = col
            current_start = row
            current_end = row

    if current_col is not None:
        if current_start == current_end:
            segments.append(f"{current_col}{current_start}")
        else:
            segments.append(f"{current_col}{current_start}:{current_col}{current_end}")

    return ",".join(segments[:10])  # Limit to avoid huge strings


def xlsx_to_sheetcompressor(filepath: str, max_rows_per_sheet: int = 500) -> str:
    """Convert an XLSX workbook to SpreadsheetLLM's SheetCompressor inverted-index format.

    Implements the three SheetCompressor modules from Dong et al. (2024):
    1. Structural anchor extraction (simplified: keep first/last rows + header neighborhoods)
    2. Inverted-index translation (group cells by value -> addresses)
    3. Data-format-aware aggregation (replace numeric values with type labels)
    """
    wb = openpyxl.load_workbook(filepath, data_only=True)
    sections = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        if ws.max_row is None or ws.max_row == 0:
            continue

        # Phase 1: Collect all cells with their addresses and values
        cells = []  # (address_str, value, row, col)
        row_count = 0
        for row_idx in range(1, (ws.max_row or 0) + 1):
            if row_count >= max_rows_per_sheet:
                break
            has_data = False
            for col_idx in range(1, (ws.max_column or 0) + 1):
                val = ws.cell(row=row_idx, column=col_idx).value
                if val is not None:
                    col_letter = get_column_letter(col_idx)
                    addr = f"{col_letter}{row_idx}"
                    cells.append((addr, val, row_idx, col_idx))
                    has_data = True
            if has_data:
                row_count += 1

        if not cells:
            continue

        # Phase 2: Inverted-index translation
        # Group cells by their string representation
        value_to_addrs: dict[str, list[str]] = {}
        for addr, val, _, _ in cells:
            val_str = _cell_value_str(val)
            if val_str not in value_to_addrs:
                value_to_addrs[val_str] = []
            value_to_addrs[val_str].append(addr)

        # Phase 3: Data-format-aware aggregation
        # For numeric values, replace with type labels and merge ranges
        dtype_to_addrs: dict[str, list[str]] = {}
        text_entries = []  # (value, merged_range) for text values

        for val_str, addrs in value_to_addrs.items():
            # Determine if this is a text value worth preserving literally
            # or a numeric value that should be aggregated by type
            sample_val = None
            for addr, val, _, _ in cells:
                if _cell_value_str(val) == val_str:
                    sample_val = val
                    break

            dtype = _classify_dtype(sample_val)

            if dtype in ("Text", "Boolean", ""):
                # Keep literal text values with inverted index
                merged = _merge_ranges(addrs)
                text_entries.append((val_str, merged))
            else:
                # Aggregate numeric/date values by type
                if dtype not in dtype_to_addrs:
                    dtype_to_addrs[dtype] = []
                dtype_to_addrs[dtype].extend(addrs)

        # Build output tuples
        tuples = []
        sheet_header = f"[Sheet: {sheet_name}]"
        tuples.append(sheet_header)

        # Text entries first (inverted index)
        for val_str, addr_range in text_entries:
            if val_str:
                tuples.append(f"({val_str}|{addr_range})")

        # Then aggregated type entries
        for dtype, addrs in sorted(dtype_to_addrs.items()):
            merged = _merge_ranges(sorted(addrs, key=lambda a: (
                re.match(r"[A-Z]+", a).group() if re.match(r"[A-Z]+", a) else "",
                int(re.search(r"\d+", a).group()) if re.search(r"\d+", a) else 0
            )))
            tuples.append(f"({dtype}|{merged})")

        sections.append("\n".join(tuples))

    return "\n\n".join(sections)


# Format registry
def xlsx_to_html(filepath: str, max_rows_per_sheet: int = 500) -> str:
    """Convert an XLSX workbook to HTML table format.

    HTML tables are consistently among the strongest performers in table-LLM
    literature (Sui et al. WSDM 2024, ImprovingAgents 2025) likely because
    LLMs see abundant HTML tables during pre-training.
    """
    wb = openpyxl.load_workbook(filepath, data_only=True)
    sections = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        html_lines = [f"<h3>{sheet_name}</h3>", "<table>"]

        row_count = 0
        first_data = True
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=True):
            if row_count >= max_rows_per_sheet:
                html_lines.append(f"<!-- showing first {max_rows_per_sheet} data rows -->")
                break
            vals = [_cell_value_str(v) for v in row]
            if all(v == "" for v in vals):
                continue

            tag = "th" if first_data else "td"
            cells = "".join(f"<{tag}>{v}</{tag}>" for v in vals)
            html_lines.append(f"  <tr>{cells}</tr>")
            first_data = False
            row_count += 1

        html_lines.append("</table>")
        sections.append("\n".join(html_lines))

    return "\n\n".join(sections)


def xlsx_to_cell_address_md(filepath: str, max_rows_per_sheet: int = 500) -> str:
    """Convert to SpreadsheetLLM's vanilla cell-address markdown format.

    From Dong et al. (SpreadsheetLLM, EMNLP 2024): each cell encoded as
    'Address,Value' pairs separated by '|', row-major order.
    E.g.: 'A1,name|B1,age\\nA2,Alice|B2,30'
    """
    wb = openpyxl.load_workbook(filepath, data_only=True)
    sections = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        lines = [f"[Sheet: {sheet_name}]"]

        row_count = 0
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=False):
            if row_count >= max_rows_per_sheet:
                lines.append(f"# showing first {max_rows_per_sheet} data rows")
                break

            cells = []
            has_data = False
            for cell in row:
                if cell.value is not None:
                    col_letter = get_column_letter(cell.column)
                    addr = f"{col_letter}{cell.row}"
                    val = _cell_value_str(cell.value)
                    cells.append(f"{addr},{val}")
                    has_data = True

            if has_data:
                lines.append("|".join(cells))
                row_count += 1

        sections.append("\n".join(lines))

    return "\n\n".join(sections)


def xlsx_to_markdown_kv(filepath: str, max_rows_per_sheet: int = 500) -> str:
    """Convert to Markdown Key-Value format (best performer in ImprovingAgents).

    Each row is a markdown section with key-value pairs. Format:
    ## Row 1
    Column1: Value1
    Column2: Value2
    """
    wb = openpyxl.load_workbook(filepath, data_only=True)
    sections = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        section_lines = [f"# Sheet: {sheet_name}"]

        # Get headers from first non-empty row
        headers = {}
        header_row = 0
        for row_idx in range(1, min(5, (ws.max_row or 1) + 1)):
            for col_idx in range(1, (ws.max_column or 0) + 1):
                val = ws.cell(row=row_idx, column=col_idx).value
                if val is not None and isinstance(val, str):
                    headers[col_idx] = val
            if len(headers) >= 2:
                header_row = row_idx
                break

        if not headers:
            # Fall back to column letters
            for col_idx in range(1, (ws.max_column or 0) + 1):
                headers[col_idx] = get_column_letter(col_idx)

        row_count = 0
        for row_idx in range(header_row + 1, (ws.max_row or 0) + 1):
            if row_count >= max_rows_per_sheet:
                section_lines.append(f"\n*showing first {max_rows_per_sheet} data rows*")
                break

            row_data = []
            has_data = False
            for col_idx in headers:
                val = ws.cell(row=row_idx, column=col_idx).value
                if val is not None:
                    has_data = True
                    row_data.append(f"{headers[col_idx]}: {_cell_value_str(val)}")

            if has_data:
                section_lines.append(f"\n## Row {row_idx}")
                section_lines.extend(row_data)
                row_count += 1

        sections.append("\n".join(section_lines))

    return "\n\n".join(sections)


def xlsx_to_raw_xlsx(filepath: str, max_rows_per_sheet: int = 500) -> str:
    """Return the raw XLSX file as base64-encoded content.

    Tests whether multimodal-capable models can parse binary spreadsheet data
    when provided as a base64 blob. The max_rows_per_sheet parameter is ignored
    since we send the entire file.
    """
    import base64
    with open(filepath, "rb") as f:
        raw = f.read()
    b64 = base64.b64encode(raw).decode("ascii")
    return f"[This is a base64-encoded Excel (.xlsx) file. Decode it to read the spreadsheet data.]\n\n{b64}"


def xlsx_to_pandas(filepath: str, max_rows_per_sheet: int = 500) -> str:
    """Pandas DataFrame .to_string() representation.

    Shows what a developer would see from `pd.read_excel(f).head(N)`.
    """
    import pandas as pd
    sections = []
    xls = pd.ExcelFile(filepath)
    for sheet_name in xls.sheet_names:
        df = pd.read_excel(xls, sheet_name=sheet_name, nrows=max_rows_per_sheet)
        sections.append(f"Sheet: {sheet_name}\n{df.to_string()}")
    return "\n\n".join(sections)


def xlsx_to_xml(filepath: str, max_rows_per_sheet: int = 500) -> str:
    """XML table format (strong performer at 56% in ImprovingAgents benchmark)."""
    wb = openpyxl.load_workbook(filepath, data_only=True)
    sections = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        xml_lines = [f'<sheet name="{sheet_name}">']

        # Get headers
        headers = {}
        for col_idx in range(1, (ws.max_column or 0) + 1):
            val = ws.cell(row=1, column=col_idx).value
            headers[col_idx] = str(val) if val else get_column_letter(col_idx)

        row_count = 0
        start_row = 2 if any(isinstance(ws.cell(row=1, column=c).value, str) for c in range(1, min(5, (ws.max_column or 1) + 1))) else 1
        for row_idx in range(start_row, (ws.max_row or 0) + 1):
            if row_count >= max_rows_per_sheet:
                xml_lines.append(f'  <!-- showing first {max_rows_per_sheet} data rows -->')
                break

            row_data = []
            has_data = False
            for col_idx in range(1, (ws.max_column or 0) + 1):
                val = ws.cell(row=row_idx, column=col_idx).value
                if val is not None:
                    has_data = True
                    tag = headers.get(col_idx, get_column_letter(col_idx))
                    # Sanitize tag name for XML
                    tag = re.sub(r'[^a-zA-Z0-9_]', '_', tag)
                    row_data.append(f'    <{tag}>{_cell_value_str(val)}</{tag}>')

            if has_data:
                xml_lines.append(f'  <row n="{row_idx}">')
                xml_lines.extend(row_data)
                xml_lines.append('  </row>')
                row_count += 1

        xml_lines.append('</sheet>')
        sections.append("\n".join(xml_lines))

    return "\n\n".join(sections)


def xlsx_to_raw_ooxml(filepath: str, max_rows_per_sheet: int = 500) -> str:
    """Extract the raw OOXML sheet XML from inside the .xlsx zip file.

    XLSX files are ZIP archives containing XML. This extracts the actual
    xl/worksheets/sheet1.xml content that Excel uses internally.
    """
    import zipfile
    sections = []
    with zipfile.ZipFile(filepath, 'r') as z:
        sheet_files = sorted(n for n in z.namelist() if n.startswith('xl/worksheets/sheet') and n.endswith('.xml'))
        for sf in sheet_files:
            content = z.read(sf).decode('utf-8')
            # Truncate if too long (raw OOXML can be huge)
            if len(content) > 200000:
                content = content[:200000] + f"\n<!-- truncated at 200KB -->"
            sections.append(f"[File: {sf}]\n{content}")
    return "\n\n".join(sections)


def xlsx_to_scf_formulas(filepath: str, max_rows_per_sheet: int = 500) -> str:
    """SCF with formula preservation (not data_only mode).

    Reads the workbook with formulas intact and encodes them in SCF format.
    Formulas appear as =FORMULA_TEXT in cell values.
    """
    wb = openpyxl.load_workbook(filepath, data_only=False)
    lines = ["#!SCF 5"]

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        safe_name = sheet_name
        if " " in sheet_name or "|" in sheet_name or '"' in sheet_name or "&" in sheet_name:
            safe_name = _scf_quote(sheet_name)
        lines.append(f"#!SHEET {safe_name}")

        header_row = None
        for row_idx in range(1, min(ws.max_row or 1, max_rows_per_sheet) + 1):
            for col_idx in range(1, (ws.max_column or 1) + 1):
                if ws.cell(row=row_idx, column=col_idx).value is not None:
                    header_row = row_idx
                    break
            if header_row:
                break
        if header_row:
            lines.append(f"#!HEADERS R{header_row}")

        row_count = 0
        for row_idx in range(1, (ws.max_row or 0) + 1):
            if row_count >= max_rows_per_sheet:
                lines.append(f"# ... showing first {max_rows_per_sheet} data rows")
                break

            cells = {}
            for col_idx in range(1, (ws.max_column or 0) + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                val = cell.value
                if val is not None:
                    # If it's a formula string (starts with =), keep it
                    if isinstance(val, str) and val.startswith("="):
                        cells[col_idx] = val  # raw formula
                    else:
                        cells[col_idx] = val

            if not cells:
                continue

            row_count += 1
            entries = []
            cursor = 1
            for col_idx in sorted(cells.keys()):
                val = cells[col_idx]
                if isinstance(val, str) and val.startswith("="):
                    scalar = val  # formula preserved as-is
                else:
                    scalar = _scf_scalar(val)
                if col_idx == cursor:
                    entries.append(scalar)
                else:
                    entries.append(f"C{col_idx}={scalar}")
                cursor = col_idx + 1

            line = f"R{row_idx} | " + " | ".join(entries) + " |"
            lines.append(line)

    return "\n".join(lines)


def xlsx_to_reverse_index_values(filepath: str, max_rows_per_sheet: int = 500) -> str:
    """Inverted-index format that preserves actual values (not type labels).

    Like SheetCompressor's inverted-index but keeps real values for all cells,
    grouping only identical text/numeric values by their cell ranges.
    JSON output format.
    """
    wb = openpyxl.load_workbook(filepath, data_only=True)
    result = {}

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        if ws.max_row is None or ws.max_row == 0:
            continue

        # Collect all cells
        value_to_addrs: dict[str, list[str]] = {}
        row_count = 0
        for row_idx in range(1, (ws.max_row or 0) + 1):
            if row_count >= max_rows_per_sheet:
                break
            has_data = False
            for col_idx in range(1, (ws.max_column or 0) + 1):
                val = ws.cell(row=row_idx, column=col_idx).value
                if val is not None:
                    val_str = _cell_value_str(val)
                    col_letter = get_column_letter(col_idx)
                    addr = f"{col_letter}{row_idx}"
                    if val_str not in value_to_addrs:
                        value_to_addrs[val_str] = []
                    value_to_addrs[val_str].append(addr)
                    has_data = True
            if has_data:
                row_count += 1

        # Merge adjacent addresses into ranges
        sheet_index = {}
        for val_str, addrs in value_to_addrs.items():
            if not val_str:
                continue
            merged = _merge_ranges(sorted(addrs, key=lambda a: (
                re.match(r"[A-Z]+", a).group() if re.match(r"[A-Z]+", a) else "",
                int(re.search(r"\d+", a).group()) if re.search(r"\d+", a) else 0
            )))
            sheet_index[val_str] = merged

        result[sheet_name] = sheet_index

    return json.dumps(result, indent=2, ensure_ascii=False)


def xlsx_to_ream_v12(filepath: str, max_rows_per_sheet: int = 500) -> str:
    """Convert an XLSX workbook to Ream format (draft 12 / wire version 11).

    New in v12 vs v9:
    - Version: #!REAM 11
    - Row spans: '12:20 | B:M=8500 |' for 2D compaction
    - Canonical vertical merge of identical horizontal segments
    """
    wb = openpyxl.load_workbook(filepath, data_only=True)
    lines = ["#!REAM 11"]

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        safe_name = sheet_name
        if " " in sheet_name or "|" in sheet_name or '"' in sheet_name or "&" in sheet_name:
            safe_name = _scf_quote(sheet_name)
        lines.append(f"#!SHEET {safe_name}")

        header_row = None
        for row_idx in range(1, min(ws.max_row or 1, max_rows_per_sheet) + 1):
            for col_idx in range(1, (ws.max_column or 1) + 1):
                if ws.cell(row=row_idx, column=col_idx).value is not None:
                    header_row = row_idx
                    break
            if header_row:
                break

        if header_row:
            lines.append(f"#!HEADERS {header_row}:{header_row}")

        # Phase 1: Build horizontal segments for each row
        # segment = (row, start_col, end_col, entries_str)
        row_segments = {}  # row -> list of (start_col, end_col, canonical_cell_text)
        row_count = 0
        for row_idx in range(1, (ws.max_row or 0) + 1):
            if row_count >= max_rows_per_sheet:
                lines.append(f"# ... showing first {max_rows_per_sheet} data rows")
                break

            cells = {}
            for col_idx in range(1, (ws.max_column or 0) + 1):
                val = ws.cell(row=row_idx, column=col_idx).value
                if val is not None:
                    cells[col_idx] = _scf_scalar(val)

            if not cells:
                continue

            row_count += 1
            # Build maximal horizontal segments of identical values
            segments = []
            sorted_cols = sorted(cells.keys())
            i = 0
            while i < len(sorted_cols):
                start_col = sorted_cols[i]
                val = cells[start_col]
                end_col = start_col
                # Extend while consecutive and identical
                while i + 1 < len(sorted_cols) and sorted_cols[i + 1] == end_col + 1 and cells[sorted_cols[i + 1]] == val:
                    i += 1
                    end_col = sorted_cols[i]
                segments.append((start_col, end_col, val))
                i += 1

            row_segments[row_idx] = segments

        # Phase 2: Vertical merge — group consecutive rows with identical segment lists
        sorted_rows = sorted(row_segments.keys())
        merged_records = []  # (start_row, end_row, segments)

        i = 0
        while i < len(sorted_rows):
            start_row = sorted_rows[i]
            segs = row_segments[start_row]
            end_row = start_row

            # Try to extend to consecutive rows with identical segments
            while i + 1 < len(sorted_rows) and sorted_rows[i + 1] == end_row + 1:
                next_row = sorted_rows[i + 1]
                if row_segments[next_row] == segs:
                    end_row = next_row
                    i += 1
                else:
                    break
            merged_records.append((start_row, end_row, segs))
            i += 1

        # Phase 3: Emit records
        for start_row, end_row, segs in merged_records:
            entries = []
            cursor = 1
            for start_col, end_col, val in segs:
                col_start_letter = get_column_letter(start_col)
                col_end_letter = get_column_letter(end_col)

                if start_col == cursor and start_col == end_col:
                    # Bare scalar at cursor
                    entries.append(val)
                elif start_col == end_col:
                    # Addressed single column
                    entries.append(f"{col_start_letter}={val}")
                elif start_col == cursor:
                    # Range starting at cursor — still need address for range
                    entries.append(f"{col_start_letter}:{col_end_letter}={val}")
                else:
                    # Addressed range
                    entries.append(f"{col_start_letter}:{col_end_letter}={val}")
                cursor = end_col + 1

            if start_row == end_row:
                prefix = str(start_row)
            else:
                prefix = f"{start_row}:{end_row}"

            line = f"{prefix} | " + " | ".join(entries) + " |"
            lines.append(line)

    return "\n".join(lines)


def xlsx_to_ream(filepath: str, max_rows_per_sheet: int = 500) -> str:
    """Convert an XLSX workbook to Ream format (draft 9 / wire version 9).

    Key design choices:
    - Version: #!REAM 9
    - Row labels are plain numbers: '12 | ...'
    - Column addresses use A1-style letters: 'B=' not 'C2='
    - Selectors use A1 notation: #!HEADERS 1:1
    - Sparse gaps use column letters: 'D=-50000'
    """
    wb = openpyxl.load_workbook(filepath, data_only=True)
    lines = ["#!REAM 9"]

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        safe_name = sheet_name
        if " " in sheet_name or "|" in sheet_name or '"' in sheet_name or "&" in sheet_name:
            safe_name = _scf_quote(sheet_name)
        lines.append(f"#!SHEET {safe_name}")

        header_row = None
        for row_idx in range(1, min(ws.max_row or 1, max_rows_per_sheet) + 1):
            for col_idx in range(1, (ws.max_column or 1) + 1):
                if ws.cell(row=row_idx, column=col_idx).value is not None:
                    header_row = row_idx
                    break
            if header_row:
                break

        if header_row:
            lines.append(f"#!HEADERS {header_row}:{header_row}")

        row_count = 0
        for row_idx in range(1, (ws.max_row or 0) + 1):
            if row_count >= max_rows_per_sheet:
                lines.append(f"# ... showing first {max_rows_per_sheet} data rows")
                break

            cells = {}
            for col_idx in range(1, (ws.max_column or 0) + 1):
                val = ws.cell(row=row_idx, column=col_idx).value
                if val is not None:
                    cells[col_idx] = val

            if not cells:
                continue

            row_count += 1
            entries = []
            cursor = 1
            for col_idx in sorted(cells.keys()):
                scalar = _scf_scalar(cells[col_idx])
                col_letter = get_column_letter(col_idx)
                if col_idx == cursor:
                    entries.append(scalar)
                else:
                    entries.append(f"{col_letter}={scalar}")
                cursor = col_idx + 1

            # Plain row number, no R prefix
            line = f"{row_idx} | " + " | ".join(entries) + " |"
            lines.append(line)

    return "\n".join(lines)


def xlsx_to_ream_addressed(filepath: str, max_rows_per_sheet: int = 500) -> str:
    """Ream format with always-on A1-style column addressing.

    Every cell gets a column-letter prefix: 'A=Revenue | B=1000 | C=1050 |'
    Uses Excel column letters that models know from training data.
    """
    wb = openpyxl.load_workbook(filepath, data_only=True)
    lines = ["#!SCF 7"]

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        safe_name = sheet_name
        if " " in sheet_name or "|" in sheet_name or '"' in sheet_name or "&" in sheet_name:
            safe_name = _scf_quote(sheet_name)
        lines.append(f"#!SHEET {safe_name}")

        header_row = None
        for row_idx in range(1, min(ws.max_row or 1, max_rows_per_sheet) + 1):
            for col_idx in range(1, (ws.max_column or 1) + 1):
                if ws.cell(row=row_idx, column=col_idx).value is not None:
                    header_row = row_idx
                    break
            if header_row:
                break

        if header_row:
            lines.append(f"#!HEADERS {header_row}:{header_row}")

        row_count = 0
        for row_idx in range(1, (ws.max_row or 0) + 1):
            if row_count >= max_rows_per_sheet:
                lines.append(f"# ... showing first {max_rows_per_sheet} data rows")
                break

            cells = {}
            for col_idx in range(1, (ws.max_column or 0) + 1):
                val = ws.cell(row=row_idx, column=col_idx).value
                if val is not None:
                    cells[col_idx] = val

            if not cells:
                continue

            row_count += 1
            entries = []
            for col_idx in sorted(cells.keys()):
                scalar = _scf_scalar(cells[col_idx])
                col_letter = get_column_letter(col_idx)
                entries.append(f"{col_letter}={scalar}")

            line = f"{row_idx} | " + " | ".join(entries) + " |"
            lines.append(line)

    return "\n".join(lines)


FORMATS = {
    "ream_v12": xlsx_to_ream_v12,
    "ream": xlsx_to_ream,
    "ream_addressed": xlsx_to_ream_addressed,
    "scf": xlsx_to_scf,
    "scf_addressed": xlsx_to_scf_addressed,
    "scf_formulas": xlsx_to_scf_formulas,
    "csv": xlsx_to_csv,
    "markdown": xlsx_to_markdown,
    "markdown_kv": xlsx_to_markdown_kv,
    "json": xlsx_to_json,
    "html": xlsx_to_html,
    "xml": xlsx_to_xml,
    "pandas": xlsx_to_pandas,
    "cell_address_md": xlsx_to_cell_address_md,
    "sheetcompressor": xlsx_to_sheetcompressor,
    "reverse_index_values": xlsx_to_reverse_index_values,
    "raw_xlsx": xlsx_to_raw_xlsx,
    "raw_ooxml": xlsx_to_raw_ooxml,
}
