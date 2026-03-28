"""
Extract questions and ground-truth answers from FRTR-Bench workbooks.
Resolves cell references in answers to actual values from the workbook.
"""

import os
import re
import json
from datetime import datetime, date
from typing import Any

import openpyxl


def _format_val(val: Any) -> str:
    """Format a cell value as a clean string for comparison."""
    if val is None:
        return ""
    if isinstance(val, bool):
        return "TRUE" if val else "FALSE"
    if isinstance(val, float):
        if val == int(val) and abs(val) < 1e15:
            return str(int(val))
        # Use enough precision but strip trailing zeros
        return f"{val:.6f}".rstrip("0").rstrip(".")
    if isinstance(val, datetime):
        return val.strftime("%Y-%m-%d")
    if isinstance(val, date):
        return val.strftime("%Y-%m-%d")
    return str(val)


def _resolve_cell(wb, ref_str: str) -> str | None:
    """Resolve 'SheetName!ColRow' to a formatted value."""
    ref_str = ref_str.strip()
    m = re.match(r"^(.+?)!([A-Z]+)(\d+)$", ref_str)
    if not m:
        return None
    sheet_name, col, row = m.group(1), m.group(2), int(m.group(3))
    # Strip quotes from sheet name
    sheet_name = sheet_name.strip("'\"")
    if sheet_name not in wb.sheetnames:
        return None
    val = wb[sheet_name][f"{col}{row}"].value
    if val is None:
        return None
    return _format_val(val)


def _try_resolve(wb, answer: str, provenance: str) -> tuple[str, str]:
    """Try to resolve the answer to a concrete value.

    Returns (resolved_value, resolution_method).
    """
    answer = answer.strip()
    prov = provenance.strip()

    # 1. "See Sheet!Cell" pattern
    m = re.match(r"^See\s+(.+?![A-Z]+\d+)", answer)
    if m:
        val = _resolve_cell(wb, m.group(1))
        if val is not None:
            return val, "cell_ref_in_answer"

    # 2. Bare "Sheet!Cell" pattern in answer
    m = re.match(r"^([A-Za-z_][\w ]*!?[A-Z]+\d+)$", answer)
    if m:
        val = _resolve_cell(wb, m.group(1))
        if val is not None:
            return val, "bare_cell_ref"

    # 3. "Max of Sheet!Range" pattern
    m = re.match(r"^Max of\s+(.+?)!([A-Z]+\d+):([A-Z]+\d+)", answer)
    if m:
        sheet_name = m.group(1).strip("'\"")
        range_str = f"{m.group(2)}:{m.group(3)}"
        if sheet_name in wb.sheetnames:
            vals = []
            for row in wb[sheet_name][range_str]:
                for cell in row:
                    if isinstance(cell.value, (int, float)):
                        vals.append(cell.value)
            if vals:
                return _format_val(max(vals)), "max_range"

    # 4. "Min of Sheet!Range" pattern
    m = re.match(r"^Min of\s+(.+?)!([A-Z]+\d+):([A-Z]+\d+)", answer)
    if m:
        sheet_name = m.group(1).strip("'\"")
        range_str = f"{m.group(2)}:{m.group(3)}"
        if sheet_name in wb.sheetnames:
            vals = []
            for row in wb[sheet_name][range_str]:
                for cell in row:
                    if isinstance(cell.value, (int, float)):
                        vals.append(cell.value)
            if vals:
                return _format_val(min(vals)), "min_range"

    # 5. Try Cell: refs in provenance
    cell_refs = re.findall(r"Cell:\s*([A-Za-z_][\w ]*![A-Z]+\d+)", prov)
    if cell_refs:
        val = _resolve_cell(wb, cell_refs[0])
        if val is not None:
            return val, "provenance_cell"

    # 6. Bare Sheet!Cell in provenance (no "Cell:" prefix)
    bare_refs = re.findall(r"([A-Za-z_][\w ]*![A-Z]+\d+)", prov)
    if bare_refs and "Image:" not in prov:
        val = _resolve_cell(wb, bare_refs[0])
        if val is not None:
            return val, "provenance_bare_ref"

    # 7. Direct text answers (Yes/No, names, etc.)
    # If answer doesn't look like a formula or ref, accept as-is
    if not re.search(r"[!=()]", answer) and not answer.startswith("See "):
        # Clean up answers like "January (use image)" → "January"
        cleaned = re.sub(r"\s*\(.*?\)\s*$", "", answer).strip()
        if cleaned:
            return cleaned, "direct_text"

    # 8. Answers with "OR" — take the simpler alternative
    if " OR " in answer:
        parts = answer.split(" OR ")
        for part in reversed(parts):  # prefer the simpler (later) part
            part = part.strip()
            # Try as cell ref
            val = _resolve_cell(wb, part)
            if val is not None:
                return val, "or_cell_ref"
            # If it's a simple string, use it
            if not re.search(r"[!=()]", part):
                return part, "or_direct"

    return answer, "unresolved"


# Visual/chart question types that require image interpretation
VISUAL_TYPES = {
    "Chart reading", "Trend reading", "Chart reading (trend)",
    "Trend/Visual", "Visual", "Visual Min", "ExtractionFromImage",
    "TrendReasoning",
}


def extract_frtr_questions(corpus_dir: str) -> list[dict]:
    """Extract all questions from FRTR-Bench workbooks."""
    questions = []
    xlsx_files = sorted(f for f in os.listdir(corpus_dir) if f.endswith(".xlsx") and f.startswith("frtr_"))

    for fname in xlsx_files:
        fpath = os.path.join(corpus_dir, fname)
        try:
            wb = openpyxl.load_workbook(fpath, data_only=True)
        except Exception as e:
            print(f"  WARN: Could not load {fname}: {e}")
            continue

        if "Questions" not in wb.sheetnames:
            print(f"  WARN: {fname} has no Questions sheet")
            continue

        qs_sheet = wb["Questions"]
        headers = {}
        for c in range(1, (qs_sheet.max_column or 0) + 1):
            h = qs_sheet.cell(row=1, column=c).value
            if h:
                headers[h.strip()] = c

        q_col = headers.get("Question", 1)
        r_col = headers.get("ReasoningType", 2)
        a_col = headers.get("Answer", 3)
        p_col = headers.get("Provenance", 4)
        d_col = headers.get("Difficulty", 5)

        # Collect non-question sheet names for context
        data_sheets = [s for s in wb.sheetnames if s not in ("Questions", "Readme")]

        for row_idx in range(2, (qs_sheet.max_row or 1) + 1):
            question_text = qs_sheet.cell(row=row_idx, column=q_col).value
            if not question_text:
                continue

            answer_raw = str(qs_sheet.cell(row=row_idx, column=a_col).value or "")
            provenance = str(qs_sheet.cell(row=row_idx, column=p_col).value or "")
            reasoning = str(qs_sheet.cell(row=row_idx, column=r_col).value or "")
            difficulty = str(qs_sheet.cell(row=row_idx, column=d_col).value or "")

            resolved_val, method = _try_resolve(wb, answer_raw, provenance)
            is_visual = reasoning in VISUAL_TYPES or "image" in provenance.lower()

            questions.append({
                "workbook": fname,
                "workbook_path": os.path.abspath(fpath),
                "data_sheets": data_sheets,
                "question": str(question_text),
                "reasoning_type": reasoning,
                "answer_raw": answer_raw,
                "answer_resolved": resolved_val,
                "resolution_method": method,
                "is_visual": is_visual,
                "provenance": provenance,
                "difficulty": difficulty,
            })

        wb.close()

    return questions


if __name__ == "__main__":
    import sys
    corpus_dir = sys.argv[1] if len(sys.argv) > 1 else "corpus/frtr"
    questions = extract_frtr_questions(corpus_dir)

    text_qs = [q for q in questions if not q["is_visual"]]
    visual_qs = [q for q in questions if q["is_visual"]]

    # Resolution stats
    methods = {}
    for q in text_qs:
        methods[q["resolution_method"]] = methods.get(q["resolution_method"], 0) + 1

    print(f"Total questions: {len(questions)}")
    print(f"  Text-answerable: {len(text_qs)}")
    print(f"  Visual/chart (excluded): {len(visual_qs)}")
    print(f"\nResolution methods (text only):")
    for m, c in sorted(methods.items(), key=lambda x: -x[1]):
        print(f"  {m}: {c}")

    # Sample resolved answers
    print(f"\n--- Sample resolved answers ---")
    for q in text_qs[:15]:
        print(f"[{q['resolution_method']}] [{q['difficulty']}] Q: {q['question'][:75]}")
        print(f"  A: {q['answer_resolved'][:80]}")
        print()

    # Write both full and text-only
    with open("data/frtr_questions_all.json", "w") as f:
        json.dump(questions, f, indent=2, default=str)
    with open("data/frtr_questions.json", "w") as f:
        json.dump(text_qs, f, indent=2, default=str)

    print(f"Saved {len(text_qs)} text questions to data/frtr_questions.json")
    print(f"Saved {len(questions)} total questions to data/frtr_questions_all.json")
